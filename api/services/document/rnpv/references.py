"""Sheet 9: References — data sources, citations, coverage summary."""

from openpyxl.styles import Font

from ._helpers import apply_header_row, set_col_widths, write_label_value
from ._styles import DARK_BLUE, MED_BLUE, NORMAL_FONT, THIN_BORDER

SN = "References"
_COVERAGE_PARAM_KEYS = ("prevalence", "diagnosed_rate", "eligible_rate", "pricing")


def _auto_collect_references(config):
    """Build a de-duped reference list from each indication's data_sources map."""
    refs = []
    seen = set()
    for ind in config.get("indications", []):
        for _, src in ind.get("data_sources", {}).items():
            if src and src not in seen:
                seen.add(src)
                refs.append(
                    {
                        "id": f"R{len(refs) + 1}",
                        "category": "Input",
                        "description": src,
                        "type": "Various",
                        "date": "",
                        "url": "",
                    }
                )
    return refs


def _write_header(ws):
    r = 1
    ws.cell(row=r, column=1, value="DATA SOURCES & REFERENCES")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 2

    headers = ["Ref #", "Category", "Description", "Type", "Date", "URL"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    return r + 1


def _write_reference_row(ws, r, ref, ref_num):
    ws.cell(row=r, column=1, value=f"[{ref.get('id', f'R{ref_num}')}]").font = Font(
        name="Calibri", size=10, bold=True, color=MED_BLUE
    )
    ws.cell(row=r, column=2, value=ref.get("category", "")).font = NORMAL_FONT
    ws.cell(row=r, column=3, value=ref.get("description", "")).font = NORMAL_FONT
    ws.cell(row=r, column=4, value=ref.get("type", "")).font = NORMAL_FONT
    ws.cell(row=r, column=5, value=ref.get("date", "")).font = NORMAL_FONT
    url = ref.get("url", "")
    c = ws.cell(row=r, column=6, value=url)
    if url and url.startswith("http"):
        c.font = Font(name="Calibri", size=10, color=MED_BLUE, underline="single")
    else:
        c.font = NORMAL_FONT
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = THIN_BORDER


def _compute_coverage(config):
    sourced = 0
    total = 0
    for ind in config.get("indications", []):
        ds = ind.get("data_sources", {})
        for p in _COVERAGE_PARAM_KEYS:
            total += 1
            if ds.get(p):
                sourced += 1
    pct = sourced / total if total > 0 else 0
    return sourced, total, pct


def build_references_sheet(wb, config, tracker):
    ws = wb.create_sheet(SN)
    ws.sheet_properties.tabColor = "00B0F0"
    set_col_widths(ws, {"A": 10, "B": 20, "C": 55, "D": 25, "E": 18, "F": 45})

    r = _write_header(ws)

    references = config.get("references") or _auto_collect_references(config)
    for idx, ref in enumerate(references, 1):
        _write_reference_row(ws, r, ref, idx)
        r += 1

    r += 2
    sourced, total, pct = _compute_coverage(config)
    write_label_value(ws, r, 1, "References", len(references))
    r += 1
    write_label_value(ws, r, 1, "Key Params Sourced", f"{sourced}/{total} ({pct:.0%})")

    config["_ref_count"] = len(references)
    config["_ref_coverage_pct"] = pct
    config["_ref_sourced"] = sourced
    config["_ref_total_params"] = total
    return ws
