"""Sheet 2: Patient Funnel — formula cascade from Assumptions to treated patients/year."""

from datetime import datetime

from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ._helpers import apply_header_row, calc_note, section_title, set_col_widths
from ._styles import (
    BOLD_FONT,
    BOTTOM_BORDER,
    DARK_BLUE,
    FORMULA_FONT,
    LIGHT_BLUE,
    LIGHT_GREEN,
    LINK_FONT,
    MED_BLUE,
    NORMAL_FONT,
    NUM_FORMAT,
    PCT2_FORMAT,
    PCT_FORMAT,
    THIN_BORDER,
    YELLOW,
)

SN = "Patient Funnel"


def _write_header(ws):
    r = 1
    ws.cell(row=r, column=1, value="PATIENT FUNNEL — FORMULA-BASED DERIVATION")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(
        ws,
        r,
        1,
        "All values are Excel formulas linking to Assumptions. Change inputs there to update.",
    )
    return r + 2


def _funnel_derivation_steps(tracker, ind_idx, geo):
    """Return the 9 (label, formula, fmt) tuples for the static funnel block."""
    p = tracker.get(f"ind{ind_idx}.{geo}.prevalence")
    d = tracker.get(f"ind{ind_idx}.{geo}.diagnosed_rate")
    e = tracker.get(f"ind{ind_idx}.{geo}.eligible_rate")
    line = tracker.get(f"ind{ind_idx}.{geo}.line_share")
    t = tracker.get(f"ind{ind_idx}.{geo}.drug_treatable_rate")
    a = tracker.get(f"ind{ind_idx}.{geo}.addressable_rate")
    addr_pts = tracker.get(f"ind{ind_idx}.{geo}.addressable")
    return [
        ("    Prevalence", f"={p}", NUM_FORMAT),
        ("    x Diagnosed Rate", f"={d}", PCT_FORMAT),
        ("    -> Diagnosed Patients", f"=INT({p}*{d})", NUM_FORMAT),
        ("    x Eligible Rate", f"={e}", PCT_FORMAT),
        ("    -> Eligible Patients", f"=INT({p}*{d}*{e})", NUM_FORMAT),
        ("    x Line Share", f"={line}", PCT_FORMAT),
        ("    x Drug-Treatable", f"={t}", PCT_FORMAT),
        ("    x Market Access", f"={a}", PCT_FORMAT),
        ("    -> Addressable Patients", f"={addr_pts}", NUM_FORMAT),
    ]


def _write_funnel_step(ws, r, label, formula, fmt):
    ws.cell(row=r, column=1, value=label).font = (
        BOLD_FONT if label.startswith("    ->") else NORMAL_FONT
    )
    ws.cell(row=r, column=1).border = THIN_BORDER
    c = ws.cell(row=r, column=2, value=formula)
    c.font = LINK_FONT
    c.number_format = fmt
    c.border = THIN_BORDER


def _write_geo_header(ws, r, geo):
    ws.cell(row=r, column=1, value=f"  Geography: {geo}").font = Font(
        name="Calibri", size=11, bold=True, color=MED_BLUE
    )


def _write_year_headers(ws, r, proj_years, base_year):
    headers = ["Year-by-Year Projection"] + [str(base_year + y) for y in range(proj_years)]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))


def _write_link_row(ws, r, label, ref_fn, proj_years, font, num_fmt):
    """Write a year-by-year row where each cell is '={ref_fn(y)}'."""
    ws.cell(row=r, column=1, value=label).font = font
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        c = ws.cell(row=r, column=2 + y, value=f"={ref_fn(y)}")
        c.font = LINK_FONT
        c.number_format = num_fmt
        c.border = THIN_BORDER


def _write_treated_row(ws, r, addr_row, pen_row, tracker, ind_idx, geo, proj_years):
    ws.cell(row=r, column=1, value="    -> Treated Patients").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=INT({cl}{addr_row}*{cl}{pen_row})")
        c.font = FORMULA_FONT
        c.number_format = NUM_FORMAT
        c.border = THIN_BORDER
        c.fill = green_fill
        tracker.set(f"funnel.ind{ind_idx}.{geo}.treated.y{y}", SN, r, col)


def _write_geo_block(ws, ind_idx, geo, tracker, proj_years, base_year, r):
    """One geography: static funnel steps + year-by-year projection (addr x pen → treated)."""
    _write_geo_header(ws, r, geo)
    r += 1

    for label, formula, fmt in _funnel_derivation_steps(tracker, ind_idx, geo):
        _write_funnel_step(ws, r, label, formula, fmt)
        r += 1
    r += 1

    _write_year_headers(ws, r, proj_years, base_year)
    r += 1

    addr_pts_ref = tracker.get(f"ind{ind_idx}.{geo}.addressable")
    _write_link_row(
        ws,
        r,
        "    Addressable Patients",
        lambda _y: addr_pts_ref,
        proj_years,
        NORMAL_FONT,
        NUM_FORMAT,
    )
    addr_row = r
    r += 1

    _write_link_row(
        ws,
        r,
        "    x Market Penetration",
        lambda y: tracker.get(f"ind{ind_idx}.pen_y{y}"),
        proj_years,
        NORMAL_FONT,
        PCT2_FORMAT,
    )
    pen_row = r
    r += 1

    _write_treated_row(ws, r, addr_row, pen_row, tracker, ind_idx, geo, proj_years)
    return r + 2


def _write_indication_total(ws, r, ind_idx, geo_keys, tracker, proj_years, ind_name):
    ws.cell(row=r, column=1, value=f"  TOTAL TREATED — {ind_name}").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    for y in range(proj_years):
        col = 2 + y
        refs = [tracker.local(f"funnel.ind{ind_idx}.{geo}.treated.y{y}") for geo in geo_keys]
        c = ws.cell(row=r, column=col, value="=" + "+".join(refs))
        c.font = FORMULA_FONT
        c.number_format = NUM_FORMAT
        c.border = THIN_BORDER
        c.fill = blue_fill
        tracker.set(f"funnel.ind{ind_idx}.total.y{y}", SN, r, col)
    for col in range(1, proj_years + 2):
        ws.cell(row=r, column=col).border = BOTTOM_BORDER


def _write_indication_section(ws, ind_idx, ind, tracker, proj_years, base_year, r):
    name = ind["name"]
    line = ind.get("line_of_therapy", "")
    title = f"INDICATION: {name}"
    if line:
        title += f"  ({line})"
    section_title(ws, r, 1, title)
    r += 1

    for geo in ind["geography_data"]:
        r = _write_geo_block(ws, ind_idx, geo, tracker, proj_years, base_year, r)

    _write_indication_total(
        ws, r, ind_idx, list(ind["geography_data"].keys()), tracker, proj_years, name
    )
    return r + 2


def _write_grand_total(ws, r, indications, tracker, proj_years):
    ws.cell(row=r, column=1, value="TOTAL TREATED — ALL INDICATIONS").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    for y in range(proj_years):
        col = 2 + y
        refs = [tracker.local(f"funnel.ind{idx}.total.y{y}") for idx in range(len(indications))]
        c = ws.cell(row=r, column=col, value="=" + "+".join(refs))
        c.font = FORMULA_FONT
        c.number_format = NUM_FORMAT
        c.border = THIN_BORDER
        c.fill = yellow_fill
        tracker.set(f"funnel.grand_total.y{y}", SN, r, col)
    return r, r + 2


def _add_total_chart(ws, grand_total_row, indications_count, proj_years, r):
    chart = LineChart()
    chart.title = "Total Treated Patients by Year"
    chart.y_axis.title = "Patients"
    chart.x_axis.title = "Year"
    chart.style = 10
    chart.width = 22
    chart.height = 12
    data_ref = Reference(
        ws,
        min_col=2,
        max_col=proj_years + 1,
        min_row=grand_total_row,
        max_row=grand_total_row,
    )
    # cats_ref was computed but never wired in the original — kept behavior identical.
    chart.add_data(data_ref, from_rows=True, titles_from_data=False)
    chart.series[0].graphicalProperties.line.width = 25000
    ws.add_chart(chart, f"A{r}")


def build_patient_funnel_sheet(wb, config, tracker):
    ws = wb.create_sheet(SN)
    ws.sheet_properties.tabColor = MED_BLUE
    set_col_widths(ws, {"A": 38, "B": 14})

    indications = config["indications"]
    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)

    r = _write_header(ws)
    for ind_idx, ind in enumerate(indications):
        r = _write_indication_section(ws, ind_idx, ind, tracker, proj_years, base_year, r)

    grand_total_row, r = _write_grand_total(ws, r, indications, tracker, proj_years)
    _add_total_chart(ws, grand_total_row, len(indications), proj_years, r)

    ws.freeze_panes = "B4"
    return ws
