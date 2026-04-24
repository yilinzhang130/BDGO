"""Sheet 1: Assumptions — single source of truth for every rNPV input."""

from datetime import datetime

from openpyxl.styles import Font, PatternFill

from ._helpers import (
    apply_header_row,
    s_curve,
    section_title,
    set_col_widths,
    write_input_cell,
)
from ._styles import (
    BOLD_FONT,
    DARK_BLUE,
    FORMULA_FONT,
    GREEN,
    INPUT_FILL,
    INPUT_FONT,
    LIGHT_BLUE,
    NORMAL_FONT,
    NUM_FORMAT,
    ORANGE,
    PCT2_FORMAT,
    PCT_FORMAT,
    RED,
    THIN_BORDER,
    USD_FORMAT,
    USD_M_FORMAT,
)

SN = "Assumptions"
_SOURCE_FONT = Font(name="Calibri", size=9, color="808080")
_SECTION_TAG_FONT = Font(name="Calibri", size=10, bold=True, color=GREEN)


def _border_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER


def _write_header(ws, meta):
    ws.cell(row=1, column=1, value=f"rNPV Valuation Model — {meta['company']} / {meta['asset']}")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    analyst = meta.get("analyst", "N/A")
    date = meta.get("date", datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=2, column=1, value=f"Analyst: {analyst}  |  Date: {date}")
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=10, color="808080")
    ws.cell(row=3, column=1, value="Color Legend:").font = Font(
        name="Calibri", size=9, bold=True, color="808080"
    )
    c = ws.cell(row=3, column=2, value="User Input (Blue)")
    c.fill = INPUT_FILL
    c.font = Font(name="Calibri", size=9, color="0000FF")
    ws.cell(row=3, column=3, value="Formula (Black)").font = Font(
        name="Calibri", size=9, color="000000"
    )
    ws.cell(row=3, column=4, value="Cross-Sheet Link (Green)").font = Font(
        name="Calibri", size=9, color="008000"
    )
    return 5


def _write_static_row(ws, row, label, value):
    ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
    ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
    _border_row(ws, row, 3)


def _write_input_row(ws, row, label, value, fmt, tracker, key, source=None):
    ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
    write_input_cell(ws, row, 2, value, fmt, tracker, key, SN)
    if source is not None:
        ws.cell(row=row, column=3, value=source).font = NORMAL_FONT
    _border_row(ws, row, 3)


def _write_general_section(ws, config, tracker, r):
    meta = config["metadata"]
    discount = config["discount"]
    costs = config["costs"]

    section_title(ws, r, 1, "GENERAL ASSUMPTIONS")
    r += 1
    for i, h in enumerate(["Parameter", "Value", "Source / Notes"], 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, 3)
    r += 1

    for label, val in [
        ("Company", meta["company"]),
        ("Asset / Drug", meta["asset"]),
        ("Modality", meta.get("modality", "N/A")),
        ("Therapeutic Area", meta.get("therapeutic_area", "N/A")),
    ]:
        _write_static_row(ws, r, label, val)
        r += 1

    _write_input_row(
        ws,
        r,
        "WACC",
        discount["wacc"],
        PCT_FORMAT,
        tracker,
        "wacc",
        source=discount.get("wacc_source", "Default biotech"),
    )
    r += 1
    _write_input_row(
        ws,
        r,
        "Tax Rate",
        discount.get("tax_rate", 0.20),
        PCT_FORMAT,
        tracker,
        "tax_rate",
        source="Global blended effective",
    )
    r += 1
    _write_input_row(
        ws,
        r,
        "Projection Years",
        discount.get("projection_years", 20),
        None,
        tracker,
        "proj_years",
    )
    r += 1
    _write_input_row(
        ws,
        r,
        "Base Year",
        meta.get("base_year", datetime.now().year),
        None,
        tracker,
        "base_year",
    )
    r += 1
    _write_input_row(
        ws,
        r,
        "COGS (% of Net Revenue)",
        costs.get("cogs_margin", 0.20),
        PCT_FORMAT,
        tracker,
        "cogs_margin",
    )
    r += 2
    return r


def _write_funnel_inputs(ws, ind, ind_idx, tracker, r):
    geo_keys = list(ind["geography_data"].keys())
    n_geos = len(geo_keys)

    for i, h in enumerate(["Patient Funnel"] + geo_keys + ["Source"], 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, n_geos + 2)
    r += 1

    params = [
        ("Prevalence / Incidence", "prevalence", NUM_FORMAT),
        ("Diagnosed Rate", "diagnosed_rate", PCT_FORMAT),
        ("Treatment Eligible Rate", "eligible_rate", PCT_FORMAT),
        ("Line-of-Therapy Share", "line_share", PCT_FORMAT),
        ("Drug-Treatable Rate", "drug_treatable_rate", PCT_FORMAT),
        ("Addressable Rate (Access)", "addressable_rate", PCT_FORMAT),
    ]
    for label, key, fmt in params:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for g_idx, geo in enumerate(geo_keys):
            val = ind["geography_data"][geo].get(key, 0)
            write_input_cell(ws, r, 2 + g_idx, val, fmt, tracker, f"ind{ind_idx}.{geo}.{key}", SN)
        source = ind.get("data_sources", {}).get(key, "")
        ws.cell(row=r, column=n_geos + 2, value=source).font = _SOURCE_FONT
        ws.cell(row=r, column=n_geos + 2).border = THIN_BORDER
        r += 1

    ws.cell(row=r, column=1, value="-> Addressable Patients").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for g_idx, geo in enumerate(geo_keys):
        col = 2 + g_idx
        refs = [
            tracker.local(f"ind{ind_idx}.{geo}.{k}")
            for k in (
                "prevalence",
                "diagnosed_rate",
                "eligible_rate",
                "line_share",
                "drug_treatable_rate",
                "addressable_rate",
            )
        ]
        formula = f"=INT({'*'.join(refs)})"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = NUM_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        tracker.set(f"ind{ind_idx}.{geo}.addressable", SN, r, col)
    ws.cell(
        row=r, column=n_geos + 2, value="Prev x Diag x Elig x Line x Treat x Access"
    ).font = Font(name="Calibri", size=9, italic=True, color="808080")
    return r + 2


def _write_pricing_block(ws, ind, ind_idx, tracker, r):
    geo_keys = list(ind["geography_data"].keys())

    ws.cell(row=r, column=1, value="PRICING").font = _SECTION_TAG_FONT
    r += 1
    for label, key, fmt in [
        ("List Price ($)", "list_price", USD_FORMAT),
        ("Gross-to-Net Ratio", "gtn", PCT_FORMAT),
    ]:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for g_idx, geo in enumerate(geo_keys):
            val = (
                ind.get("pricing", {}).get(geo, 0)
                if key == "list_price"
                else ind.get("gross_to_net", {}).get(geo, 0.70)
            )
            write_input_cell(ws, r, 2 + g_idx, val, fmt, tracker, f"ind{ind_idx}.{geo}.{key}", SN)
        r += 1

    ws.cell(row=r, column=1, value="-> Net Price ($)").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for g_idx, geo in enumerate(geo_keys):
        col = 2 + g_idx
        lp = tracker.local(f"ind{ind_idx}.{geo}.list_price")
        gtn = tracker.local(f"ind{ind_idx}.{geo}.gtn")
        c = ws.cell(row=r, column=col, value=f"={lp}*{gtn}")
        c.font = FORMULA_FONT
        c.number_format = USD_FORMAT
        c.border = THIN_BORDER
        tracker.set(f"ind{ind_idx}.{geo}.net_price", SN, r, col)
    return r + 2


def _compute_penetration_values(peak_pen, ramp_yrs, loe_year, post_loe, launch_offset, proj_years):
    vals = []
    for yr_offset in range(proj_years):
        ysl = yr_offset - launch_offset
        if ysl < 0:
            vals.append(0)
        elif ysl < loe_year:
            vals.append(s_curve(ysl + 1, peak_pen, ramp_yrs))
        else:
            years_post_loe = ysl - loe_year
            base_pen = s_curve(loe_year, peak_pen, ramp_yrs)
            vals.append(base_pen * ((1 - post_loe) ** (years_post_loe + 1)))
    return vals


def _write_penetration_block(ws, ind, ind_idx, config, tracker, r):
    ws.cell(row=r, column=1, value="MARKET PENETRATION").font = _SECTION_TAG_FONT
    r += 1

    pen = ind.get("penetration_curve", {})
    peak_pen = pen.get("peak", 0.15)
    ramp_yrs = pen.get("ramp_years", 7)
    loe_year = pen.get("loe_year_from_launch", 12)
    post_loe = pen.get("post_loe_erosion_per_year", 0.30)
    launch_offset = ind.get("years_to_launch", 5)

    params = [
        ("Peak Penetration", peak_pen, PCT_FORMAT),
        ("Ramp-up Years", ramp_yrs, None),
        ("LOE Year (from launch)", loe_year, None),
        ("Post-LOE Erosion/yr", post_loe, PCT_FORMAT),
        ("Years to Launch", launch_offset, None),
    ]
    for label, val, fmt in params:
        ws.cell(row=r, column=1, value=f"  {label}").font = NORMAL_FONT
        key = f"ind{ind_idx}.pen.{label.lower().replace(' ', '_').replace('/', '_')}"
        write_input_cell(ws, r, 2, val, fmt, tracker, key, SN)
        r += 1

    tracker.set(f"ind{ind_idx}.years_to_launch", SN, r - 1, 2)

    r += 1
    ws.cell(row=r, column=1, value="Penetration by Year:").font = _SECTION_TAG_FONT
    r += 1

    proj_years = config["discount"].get("projection_years", 20)
    base_year = config["metadata"].get("base_year", datetime.now().year)
    year_headers = ["Year"] + [str(base_year + y) for y in range(proj_years)]
    for i, h in enumerate(year_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(year_headers))
    r += 1

    pen_vals = _compute_penetration_values(
        peak_pen, ramp_yrs, loe_year, post_loe, launch_offset, proj_years
    )
    ws.cell(row=r, column=1, value="Penetration %").font = NORMAL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        write_input_cell(
            ws, r, 2 + y, pen_vals[y], PCT2_FORMAT, tracker, f"ind{ind_idx}.pen_y{y}", SN
        )
    return r + 2


def _write_pos_block(ws, ind, ind_idx, tracker, r):
    ws.cell(row=r, column=1, value="PROBABILITY OF SUCCESS").font = _SECTION_TAG_FONT
    r += 1
    pos = ind.get("pos", {})
    ws.cell(row=r, column=1, value="  Current Phase").font = NORMAL_FONT
    ws.cell(row=r, column=2, value=pos.get("current_phase", "Phase 1")).font = INPUT_FONT
    r += 1

    for trans_name, trans_val in pos.get("phase_transitions", {}).items():
        label = trans_name.replace("_to_", " -> ").replace("_", " ").title()
        ws.cell(row=r, column=1, value=f"  {label}").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, trans_val, PCT_FORMAT, tracker, f"ind{ind_idx}.pos.{trans_name}", SN
        )
        r += 1

    cum_pos = pos.get("cumulative", 0.10)
    ws.cell(row=r, column=1, value="  -> Cumulative PoS").font = BOLD_FONT
    write_input_cell(ws, r, 2, cum_pos, PCT_FORMAT, tracker, f"ind{ind_idx}.cum_pos", SN)
    ws.cell(row=r, column=2).font = Font(name="Calibri", size=12, bold=True, color=RED)
    return r + 2


def _write_indication(ws, ind, ind_idx, config, tracker, r):
    section_title(ws, r, 1, f"INDICATION {ind_idx + 1}: {ind['name']}")
    if ind.get("line_of_therapy"):
        ws.cell(row=r, column=3, value=f"Line: {ind['line_of_therapy']}")
        ws.cell(row=r, column=3).font = Font(name="Calibri", size=10, bold=True, color=ORANGE)
    r += 1

    r = _write_funnel_inputs(ws, ind, ind_idx, tracker, r)
    r = _write_pricing_block(ws, ind, ind_idx, tracker, r)
    r = _write_penetration_block(ws, ind, ind_idx, config, tracker, r)
    r = _write_pos_block(ws, ind, ind_idx, tracker, r)
    return r


def _write_rd_costs(ws, config, tracker, r):
    costs = config["costs"]
    section_title(ws, r, 1, "R&D COST ASSUMPTIONS ($M)")
    r += 1
    rd_headers = [
        "Phase",
        "Total Cost ($M)",
        "Duration (Yr)",
        "Start Year (offset)",
        "# Trials",
        "Patients/Trial",
        "# Sites",
        "Source",
    ]
    for i, h in enumerate(rd_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(rd_headers))
    r += 1

    for p_idx, phase in enumerate(costs.get("rd_by_phase", [])):
        ws.cell(row=r, column=1, value=phase.get("phase", "")).font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        write_input_cell(
            ws, r, 2, phase.get("cost_mm", 0), USD_M_FORMAT, tracker, f"rd{p_idx}.cost", SN
        )
        write_input_cell(
            ws, r, 3, phase.get("duration_years", 1), None, tracker, f"rd{p_idx}.duration", SN
        )
        write_input_cell(
            ws, r, 4, phase.get("start_year", 0), None, tracker, f"rd{p_idx}.start", SN
        )
        write_input_cell(ws, r, 5, phase.get("num_trials", 1), None)
        write_input_cell(ws, r, 6, phase.get("patients_per_trial", ""), None)
        write_input_cell(ws, r, 7, phase.get("num_sites", ""), None)
        ws.cell(row=r, column=8, value=phase.get("source", "")).font = _SOURCE_FONT
        ws.cell(row=r, column=8).border = THIN_BORDER
        r += 1

    config["_rd_phase_count"] = len(costs.get("rd_by_phase", []))
    r += 1

    cmc = costs.get("cmc", {})
    if cmc:
        cmc_total = sum(v for v in cmc.values() if isinstance(v, (int, float)))
        ws.cell(row=r, column=1, value="CMC / Manufacturing Total ($M)").font = NORMAL_FONT
        write_input_cell(ws, r, 2, cmc_total, USD_M_FORMAT, tracker, "cmc_total", SN)
        ws.cell(row=r, column=3, value="Spread over 5 years").font = _SOURCE_FONT
        r += 2

    return r


def _write_sga_sales(ws, sales, tracker, r):
    ramp = sales.get("ramp_schedule", [0.3, 0.6, 1.0])
    rows = [
        ("Sales Reps (Full Deploy)", sales.get("reps", 100), None, "sga.reps"),
        ("Cost per Rep ($K/yr)", sales.get("cost_per_rep_k", 280), USD_FORMAT, "sga.cost_per_rep"),
        ("Ramp: Pre-launch %", ramp[0] if len(ramp) > 0 else 0.3, PCT_FORMAT, "sga.ramp0"),
        ("Ramp: Launch %", ramp[1] if len(ramp) > 1 else 0.6, PCT_FORMAT, "sga.ramp1"),
        ("Ramp: Year 2+ %", ramp[2] if len(ramp) > 2 else 1.0, PCT_FORMAT, "sga.ramp2"),
    ]
    for label, val, fmt, key in rows:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        write_input_cell(ws, r, 2, val, fmt, tracker, key, SN)
        r += 1
    return r


def _write_sga_msls(ws, msls, tracker, r):
    rows = [
        ("MSL Count", msls.get("count", 20), None, "sga.msl_count"),
        ("Cost per MSL ($K/yr)", msls.get("cost_per_msl_k", 350), USD_FORMAT, "sga.msl_cost"),
    ]
    for label, val, fmt, key in rows:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        write_input_cell(ws, r, 2, val, fmt, tracker, key, SN)
        r += 1
    return r


def _write_sga_marketing(ws, mktg, tracker, r):
    rows = [
        ("Congress/KOL ($M/yr)", mktg.get("congress_annual_mm", 3), USD_M_FORMAT, "sga.congress"),
        ("Publications ($M/yr)", mktg.get("publications_mm", 1), USD_M_FORMAT, "sga.pubs"),
        (
            "Digital Marketing ($M/yr)",
            mktg.get("digital_marketing_mm", 2),
            USD_M_FORMAT,
            "sga.digital",
        ),
        (
            "Pre-Launch Mktg Total ($M)",
            mktg.get("prelaunch_total_mm", 5),
            USD_M_FORMAT,
            "sga.prelaunch",
        ),
    ]
    for label, val, fmt, key in rows:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        write_input_cell(ws, r, 2, val, fmt, tracker, key, SN)
        r += 1
    return r


def _write_sga_section(ws, config, tracker, r):
    costs = config["costs"]
    section_title(ws, r, 1, "SG&A ASSUMPTIONS")
    r += 1
    sga = costs.get("sga", {})

    sales = sga.get("sales_team", {})
    if sales:
        r = _write_sga_sales(ws, sales, tracker, r)

    msls = sga.get("msls", {})
    if msls:
        r = _write_sga_msls(ws, msls, tracker, r)

    mktg = sga.get("marketing", {})
    if mktg:
        r = _write_sga_marketing(ws, mktg, tracker, r)

    ws.cell(row=r, column=1, value="G&A (% of Revenue)").font = NORMAL_FONT
    write_input_cell(
        ws, r, 2, sga.get("ga_pct_of_revenue", 0.05), PCT_FORMAT, tracker, "sga.ga_pct", SN
    )
    return r + 1


def build_assumptions_sheet(wb, config, tracker):
    ws = wb.active
    ws.title = SN
    ws.sheet_properties.tabColor = DARK_BLUE
    set_col_widths(ws, {"A": 38, "B": 18, "C": 22, "D": 18, "E": 18, "F": 18, "G": 18})

    r = _write_header(ws, config["metadata"])
    r = _write_general_section(ws, config, tracker, r)
    for ind_idx, ind in enumerate(config["indications"]):
        r = _write_indication(ws, ind, ind_idx, config, tracker, r)
    r = _write_rd_costs(ws, config, tracker, r)
    _write_sga_section(ws, config, tracker, r)

    ws.freeze_panes = "A5"
    return ws
