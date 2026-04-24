"""Sheet 6: rNPV Model — risk-adjusted NPV + Python-side mirror for downstream sheets.

The Excel side writes formulas so users can edit Assumptions and see live rNPV.
The Python side re-computes identical numbers (mid-year discount, mirrored SG&A
timing) so QC / Summary / Sensitivity sheets can read actual floats from
`config["_npv"]`, `config["_fcf_vals"]`, etc.
"""

import math
from datetime import datetime

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ._helpers import (
    apply_header_row,
    calc_note,
    s_curve,
    section_title,
    set_col_widths,
    write_input_cell,
)
from ._styles import (
    BOLD_FONT,
    DARK_BLUE,
    FORMULA_FONT,
    GREEN,
    LIGHT_BLUE,
    LIGHT_GREEN,
    LINK_FONT,
    NORMAL_FONT,
    NUM_FORMAT,
    PCT_FORMAT,
    RED,
    THIN_BORDER,
    USD_M_FORMAT,
    YELLOW,
)

SN = "rNPV Model"


def _write_header(ws, config, proj_years, base_year):
    r = 1
    ws.cell(row=r, column=1, value="rNPV MODEL — FORMULA-BASED")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(ws, r, 1, "rNPV = SUM[FCF(t) x PoS(t) x 1/(1+WACC)^(t+0.5)]  (mid-year convention)")
    r += 2

    npv_col = proj_years + 2
    headers = ["($M)"] + [str(base_year + y) for y in range(proj_years)] + ["NPV"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    return r + 1, npv_col


def _write_fcf_link_row(ws, tracker, proj_years, r):
    ws.cell(row=r, column=1, value="Unrisked FCF").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=f"={tracker.get(f'pl.fcf.y{y}')}")
        c.font = LINK_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    return r + 2


def _write_per_indication_pos(ws, indications, tracker, r):
    for ind_idx, ind in enumerate(indications):
        cum_pos_ref = tracker.get(f"ind{ind_idx}.cum_pos")
        ws.cell(row=r, column=1, value=f"  {ind['name']} Cum PoS:").font = NORMAL_FONT
        c = ws.cell(row=r, column=2, value=f"={cum_pos_ref}")
        c.font = LINK_FONT
        c.number_format = PCT_FORMAT
        c.border = THIN_BORDER
        r += 1
    return r + 1


def _write_pos_timeline(ws, tracker, proj_years, first_launch, total_cum_pos, r):
    ws.cell(row=r, column=1, value="PoS (by year)").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        pos_val = total_cum_pos if y < first_launch else 1.0
        write_input_cell(ws, r, 2 + y, pos_val, PCT_FORMAT, tracker, f"rnpv.pos.y{y}", SN)
    pos_row = r
    r += 1
    calc_note(ws, r, 1, f"Pre-approval: {total_cum_pos:.1%} | Post-approval: 100%")
    return r + 2, pos_row


def _write_formula_row(ws, r, label, formula_fn, proj_years, fill=None):
    ws.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=formula_fn(y))
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        if fill is not None:
            c.fill = fill


def _write_npv_cell(ws, r, npv_col, proj_years, val_font):
    end = get_column_letter(proj_years + 1)
    formula = f"=SUM({get_column_letter(2)}{r}:{end}{r})"
    c = ws.cell(row=r, column=npv_col, value=formula)
    c.font = val_font
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER


def _write_risked_and_discount(ws, tracker, proj_years, npv_col, fcf_row, pos_row, r):
    # Risk-Adjusted FCF
    green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    _write_formula_row(
        ws,
        r,
        "Risk-Adjusted FCF",
        lambda y: f"={get_column_letter(2 + y)}{fcf_row}*{get_column_letter(2 + y)}{pos_row}",
        proj_years,
        fill=green_fill,
    )
    risked_row = r
    r += 2

    # Discount Factor
    wacc_ref = tracker.get("wacc")
    ws.cell(row=r, column=1, value="Discount Factor").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=f"=1/(1+{wacc_ref})^({y}+0.5)")
        c.font = FORMULA_FONT
        c.number_format = "0.0000"
        c.border = THIN_BORDER
    df_row = r
    r += 1

    # PV of Risked FCF + NPV SUM
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    _write_formula_row(
        ws,
        r,
        "PV of Risk-Adj FCF",
        lambda y: f"={get_column_letter(2 + y)}{risked_row}*{get_column_letter(2 + y)}{df_row}",
        proj_years,
        fill=yellow_fill,
    )
    _write_npv_cell(ws, r, npv_col, proj_years, Font(name="Calibri", size=14, bold=True, color=RED))
    tracker.set("rnpv.npv", SN, r, npv_col)
    r += 2

    # Unrisked PV
    ws.cell(row=r, column=1, value="Unrisked PV of FCF").font = NORMAL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"={cl}{fcf_row}*{cl}{df_row}")
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
    _write_npv_cell(
        ws, r, npv_col, proj_years, Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
    )
    tracker.set("rnpv.unrisked_npv", SN, r, npv_col)
    return r + 2, wacc_ref


def _write_summary(ws, config, tracker, wacc_ref, r):
    section_title(ws, r, 1, "VALUATION SUMMARY")
    r += 1
    npv_ref = tracker.local("rnpv.npv")
    unrisked_ref = tracker.local("rnpv.unrisked_npv")

    summary_items = [
        ("rNPV ($M)", f"={npv_ref}", USD_M_FORMAT),
        ("Unrisked NPV ($M)", f"={unrisked_ref}", USD_M_FORMAT),
        ("Risk Discount (rNPV/uNPV)", f"=IFERROR({npv_ref}/{unrisked_ref},0)", PCT_FORMAT),
        ("WACC", f"={wacc_ref}", PCT_FORMAT),
    ]
    for label, formula, fmt in summary_items:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        c = ws.cell(row=r, column=2, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
        c.number_format = fmt
        c.border = THIN_BORDER
        r += 1

    shares = config.get("metadata", {}).get("shares_outstanding_mm")
    if shares and shares > 0:
        ws.cell(row=r, column=1, value="Shares Outstanding (M)").font = BOLD_FONT
        ws.cell(row=r, column=2, value=shares).number_format = NUM_FORMAT
        r += 1
        ws.cell(row=r, column=1, value="rNPV Per Share ($)").font = BOLD_FONT
        c = ws.cell(row=r, column=2, value=f"={npv_ref}/{shares}")
        c.font = Font(name="Calibri", size=16, bold=True, color=GREEN)
        c.number_format = "$#,##0.00"


# ────────────────────────────────────────────────────────────────
# Python-side mirror (QC / Summary / Sensitivity consumers)
# ────────────────────────────────────────────────────────────────


def _penetration_value(ysl, peak_pen, ramp_yrs, loe_year, post_loe):
    if ysl < 0:
        return 0
    if ysl < loe_year:
        return s_curve(ysl + 1, peak_pen, ramp_yrs)
    base_p = s_curve(loe_year, peak_pen, ramp_yrs)
    return base_p * ((1 - post_loe) ** (ysl - loe_year + 1))


def _indication_revenue_by_year(ind, proj_years):
    """Return {year: revenue_in_mm} for one indication, summed across geos."""
    pen = ind.get("penetration_curve", {})
    peak_pen = pen.get("peak", 0.15)
    ramp_yrs = pen.get("ramp_years", 7)
    loe_year = pen.get("loe_year_from_launch", 12)
    post_loe = pen.get("post_loe_erosion_per_year", 0.30)
    launch_offset = ind.get("years_to_launch", 5)

    yearly = {}
    for geo, gd in ind["geography_data"].items():
        addr = gd.get("prevalence", 0)
        for k in (
            "diagnosed_rate",
            "eligible_rate",
            "line_share",
            "drug_treatable_rate",
            "addressable_rate",
        ):
            addr *= gd.get(k, 1)
        net_price = ind.get("pricing", {}).get(geo, 0) * ind.get("gross_to_net", {}).get(geo, 0.7)
        for y in range(proj_years):
            pv = _penetration_value(y - launch_offset, peak_pen, ramp_yrs, loe_year, post_loe)
            treated = int(addr * pv)
            yearly[y] = yearly.get(y, 0) + treated * net_price / 1e6
    return yearly


def _rd_cost_year(y, rd_phases, cmc_total):
    rd = sum(
        p.get("cost_mm", 0) / max(p.get("duration_years", 1), 0.5)
        for p in rd_phases
        if p.get("start_year", 0)
        <= y
        < p.get("start_year", 0) + math.ceil(max(p.get("duration_years", 1), 0.5))
    )
    if y < 5:
        rd += cmc_total / 5
    return rd


def _sga_sales_year(yfl, sales_cfg):
    reps = sales_cfg.get("reps", 0)
    cpr = sales_cfg.get("cost_per_rep_k", 0)
    ramp = sales_cfg.get("ramp_schedule", [0.3, 0.6, 1.0])
    if yfl == -1:
        mult = ramp[0] if ramp else 0.3
    elif yfl == 0:
        mult = ramp[1] if len(ramp) > 1 else ramp[-1]
    elif yfl >= 1:
        mult = ramp[2] if len(ramp) > 2 else ramp[-1]
    else:
        return 0
    return reps * cpr / 1000 * mult


def _sga_msl_year(yfl, msls_cfg):
    if yfl < -2:
        return 0
    base = msls_cfg.get("count", 0) * msls_cfg.get("cost_per_msl_k", 0) / 1000
    if base == 0:
        return 0
    return base * 0.5 if yfl < 0 else base


def _sga_marketing_year(yfl, mktg_cfg):
    if yfl in (-2, -1):
        return mktg_cfg.get("prelaunch_total_mm", 0) / 2
    if yfl >= 0:
        return (
            mktg_cfg.get("congress_annual_mm", 0)
            + mktg_cfg.get("publications_mm", 0)
            + mktg_cfg.get("digital_marketing_mm", 0)
        )
    return 0


def _sga_ga_year(yfl, rev, ga_pct):
    return max(rev * ga_pct, 2) if -3 <= yfl <= 0 else rev * ga_pct


def _fcf_year(y, rev, config, first_launch, tax_rate):
    costs = config["costs"]
    sga_cfg = costs.get("sga", {})
    rd = _rd_cost_year(
        y,
        costs.get("rd_by_phase", []),
        sum(v for v in costs.get("cmc", {}).values() if isinstance(v, (int, float))),
    )
    cogs = rev * costs.get("cogs_margin", 0.20)
    yfl = y - first_launch
    sga = (
        _sga_sales_year(yfl, sga_cfg.get("sales_team", {}))
        + _sga_msl_year(yfl, sga_cfg.get("msls", {}))
        + _sga_marketing_year(yfl, sga_cfg.get("marketing", {}))
        + _sga_ga_year(yfl, rev, sga_cfg.get("ga_pct_of_revenue", 0.05))
    )
    ebit = rev - cogs - rd - sga
    tax = max(0, ebit) * tax_rate
    return ebit - tax


def _compute_mirror(config, proj_years, first_launch, total_cum_pos):
    """Compute Python-side rev_vals / fcf_vals / NPV to match Excel formulas."""
    indications = config["indications"]
    wacc_val = config["discount"]["wacc"]
    tax_rate = config["discount"].get("tax_rate", 0.20)

    # Per-year revenue summed across indications
    rev_by_year = {}
    for ind in indications:
        for y, rev in _indication_revenue_by_year(ind, proj_years).items():
            rev_by_year[y] = rev_by_year.get(y, 0) + rev
    rev_vals = [rev_by_year.get(y, 0) for y in range(proj_years)]

    fcf_vals = [
        _fcf_year(y, rev_vals[y], config, first_launch, tax_rate) for y in range(proj_years)
    ]
    pos_tl = [total_cum_pos if y < first_launch else 1.0 for y in range(proj_years)]
    disc = [1 / ((1 + wacc_val) ** (y + 0.5)) for y in range(proj_years)]

    npv = sum(fcf_vals[y] * pos_tl[y] * disc[y] for y in range(proj_years))
    unrisked_npv = sum(fcf_vals[y] * disc[y] for y in range(proj_years))
    peak_rev = max(rev_vals) if rev_vals else 0
    peak_rev_year = (
        config["metadata"].get("base_year", datetime.now().year) + rev_vals.index(peak_rev)
        if peak_rev > 0
        else "N/A"
    )

    return {
        "rev_vals": rev_vals,
        "fcf_vals": fcf_vals,
        "pos_tl": pos_tl,
        "disc": disc,
        "npv": npv,
        "unrisked_npv": unrisked_npv,
        "peak_rev": peak_rev,
        "peak_rev_year": peak_rev_year,
    }


def _store_mirror(config, tracker, mirror):
    config["_npv_formula_ref"] = tracker.get("rnpv.npv")
    config["_unrisked_formula_ref"] = tracker.get("rnpv.unrisked_npv")
    config["_pos_timeline"] = mirror["pos_tl"]
    config["_npv"] = mirror["npv"]
    config["_unrisked_npv"] = mirror["unrisked_npv"]
    config["_peak_rev"] = mirror["peak_rev"]
    config["_peak_rev_year"] = mirror["peak_rev_year"]
    config["_rev_vals"] = mirror["rev_vals"]
    config["_fcf_vals"] = mirror["fcf_vals"]
    config["_disc_factors"] = mirror["disc"]
    config["_risked_fcf"] = [
        mirror["fcf_vals"][y] * mirror["pos_tl"][y] for y in range(len(mirror["fcf_vals"]))
    ]
    config["_computed_revenues"] = dict(enumerate(mirror["rev_vals"]))
    config["_computed_costs"] = {"rd": [], "cogs": [], "sga": []}


def build_rnpv_sheet(wb, config, tracker):
    ws = wb.create_sheet(SN)
    ws.sheet_properties.tabColor = RED
    set_col_widths(ws, {"A": 40})

    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)
    indications = config["indications"]
    first_launch = min(ind.get("years_to_launch", 5) for ind in indications)
    total_cum_pos = max(ind.get("pos", {}).get("cumulative", 0.10) for ind in indications)

    r, npv_col = _write_header(ws, config, proj_years, base_year)
    fcf_row = r
    r = _write_fcf_link_row(ws, tracker, proj_years, r)

    section_title(ws, r, 1, "PROBABILITY OF SUCCESS (by year)")
    r += 1
    r = _write_per_indication_pos(ws, indications, tracker, r)
    r, pos_row = _write_pos_timeline(ws, tracker, proj_years, first_launch, total_cum_pos, r)

    r, wacc_ref = _write_risked_and_discount(ws, tracker, proj_years, npv_col, fcf_row, pos_row, r)
    _write_summary(ws, config, tracker, wacc_ref, r)

    ws.freeze_panes = "B5"

    mirror = _compute_mirror(config, proj_years, first_launch, total_cum_pos)
    _store_mirror(config, tracker, mirror)
    return ws
