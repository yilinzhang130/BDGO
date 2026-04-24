"""Sheet 7: Sensitivity — tornado + scenario analysis against the Python-mirror rNPV."""

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

from ._helpers import apply_header_row, section_title, set_col_widths
from ._styles import (
    BOLD_FONT,
    DARK_BLUE,
    GREEN,
    NORMAL_FONT,
    PCT_FORMAT,
    RED,
    THIN_BORDER,
    USD_M_FORMAT,
)

SN = "Sensitivity"


def _make_quick_rnpv(config):
    """Closure that recomputes rNPV under revenue / cost / WACC / PoS multiplier overrides.

    Mirrors the math in rnpv_sheet._compute_mirror (mid-year discount, PoS 100%
    post-launch) so tornado/scenario swings line up with the base NPV that the
    rNPV sheet published.
    """
    wacc = config["discount"]["wacc"]
    proj_years = config["discount"].get("projection_years", 20)
    indications = config["indications"]
    fcf_vals = config["_fcf_vals"]
    base_cum_pos = max(ind.get("pos", {}).get("cumulative", 0.10) for ind in indications)
    first_launch = min(ind.get("years_to_launch", 5) for ind in indications)

    def quick_rnpv(rev_mult=1.0, cost_mult=1.0, wacc_override=None, pos_mult=1.0):
        w = wacc_override if wacc_override else wacc
        adj_pos = min(base_cum_pos * pos_mult, 1.0)
        total = 0.0
        for y in range(proj_years):
            fcf = fcf_vals[y] * rev_mult if fcf_vals[y] > 0 else fcf_vals[y] * cost_mult
            pos = adj_pos if y < first_launch else 1.0
            df = 1 / ((1 + w) ** (y + 0.5))
            total += fcf * pos * df
        return total

    return quick_rnpv


def _tornado_variables(quick_rnpv, wacc):
    """Return the 4 tornado inputs: (label, low_in, base_in, high_in, calc_fn, fmt).

    `fmt` is the format string applied to low/base/high for display (`.0%` for
    multipliers, `.1%` for WACC). `calc_fn(input)` returns rNPV.
    """
    return [
        ("Peak Revenue (+-30%)", 0.70, 1.00, 1.30, lambda v: quick_rnpv(rev_mult=v), ".0%"),
        ("Costs (+-30%)", 0.70, 1.00, 1.30, lambda v: quick_rnpv(cost_mult=v), ".0%"),
        ("PoS (+-50%)", 0.50, 1.00, 1.50, lambda v: quick_rnpv(pos_mult=v), ".0%"),
        (
            "WACC (+-3pp)",
            wacc + 0.03,
            wacc,
            wacc - 0.03,
            lambda v: quick_rnpv(wacc_override=v),
            ".1%",
        ),
    ]


def _compute_tornado_data(quick_rnpv, wacc):
    """Return sorted list of (label, lo_str, base_str, hi_str, npv_lo, npv_hi, swing)."""
    rows = []
    for label, lo, base, hi, calc_fn, fmt in _tornado_variables(quick_rnpv, wacc):
        npv_lo = calc_fn(lo)
        npv_hi = calc_fn(hi)
        # Cost multiplier flips the direction: higher cost → lower NPV.
        if "Cost" in label:
            npv_lo, npv_hi = npv_hi, npv_lo
        rows.append(
            (
                label,
                format(lo, fmt),
                format(base, fmt),
                format(hi, fmt),
                npv_lo,
                npv_hi,
                abs(npv_hi - npv_lo),
            )
        )
    rows.sort(key=lambda x: x[6], reverse=True)
    return rows


def _write_tornado_row(ws, r, item):
    label, lo, base, hi, npv_lo, npv_hi, swing = item
    ws.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws.cell(row=r, column=2, value=lo).font = NORMAL_FONT
    ws.cell(row=r, column=3, value=base).font = NORMAL_FONT
    ws.cell(row=r, column=4, value=hi).font = NORMAL_FONT
    ws.cell(row=r, column=5, value=npv_lo).number_format = USD_M_FORMAT
    ws.cell(row=r, column=6, value=npv_hi).number_format = USD_M_FORMAT
    ws.cell(row=r, column=7, value=swing).font = BOLD_FONT
    ws.cell(row=r, column=7).number_format = USD_M_FORMAT
    for col in range(1, 8):
        ws.cell(row=r, column=col).border = THIN_BORDER


def _write_tornado_section(ws, quick_rnpv, wacc, r):
    section_title(ws, r, 1, "TORNADO ANALYSIS")
    r += 1
    headers = ["Variable", "Low", "Base", "High", "rNPV Low ($M)", "rNPV High ($M)", "Swing ($M)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    tornado_data = _compute_tornado_data(quick_rnpv, wacc)
    tornado_start = r
    for item in tornado_data:
        _write_tornado_row(ws, r, item)
        r += 1
    return r + 1, tornado_start, r - 1


def _add_tornado_chart(ws, tornado_start, tornado_end, r):
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Tornado — rNPV Impact ($M)"
    chart.style = 10
    chart.width = 22
    chart.height = 14
    data_low = Reference(ws, min_col=5, min_row=tornado_start - 1, max_row=tornado_end)
    data_high = Reference(ws, min_col=6, min_row=tornado_start - 1, max_row=tornado_end)
    cats = Reference(ws, min_col=1, min_row=tornado_start, max_row=tornado_end)
    chart.add_data(data_low, titles_from_data=True)
    chart.add_data(data_high, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = RED
    chart.series[1].graphicalProperties.solidFill = GREEN
    ws.add_chart(chart, f"A{r}")
    return r + 18


def _scenario_rows(config):
    scenarios = config.get("scenarios", {})
    bull = scenarios.get("bull", {})
    bear = scenarios.get("bear", {})
    return [
        ("Bear", bear.get("revenue_multiplier", 0.7), bear.get("pos_multiplier", 0.7)),
        ("Base", 1.0, 1.0),
        ("Bull", bull.get("revenue_multiplier", 1.3), bull.get("pos_multiplier", 1.3)),
    ]


def _write_scenarios(ws, quick_rnpv, config, base_npv, r):
    section_title(ws, r, 1, "SCENARIO COMPARISON")
    r += 1
    headers = ["Scenario", "Rev Mult", "PoS Mult", "rNPV ($M)", "vs Base"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    for label, rm, pm in _scenario_rows(config):
        val = quick_rnpv(rev_mult=rm, pos_mult=pm)
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=2, value=rm).number_format = PCT_FORMAT
        ws.cell(row=r, column=3, value=pm).number_format = PCT_FORMAT
        ws.cell(row=r, column=4, value=val).number_format = USD_M_FORMAT
        ws.cell(row=r, column=5, value=val - base_npv).number_format = USD_M_FORMAT
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1
    return r


def build_sensitivity_sheet(wb, config, tracker):
    ws = wb.create_sheet(SN)
    ws.sheet_properties.tabColor = "C00000"
    set_col_widths(ws, {"A": 32, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})

    r = 1
    ws.cell(row=r, column=1, value="SENSITIVITY ANALYSIS")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 2

    quick_rnpv = _make_quick_rnpv(config)
    wacc = config["discount"]["wacc"]

    r, tornado_start, tornado_end = _write_tornado_section(ws, quick_rnpv, wacc, r)
    r = _add_tornado_chart(ws, tornado_start, tornado_end, r)
    _write_scenarios(ws, quick_rnpv, config, config["_npv"], r)

    ws.freeze_panes = "A3"
    return ws
