"""Sheet 5: P&L & Cash Flow — formula cascade from Revenue + Cost sheets."""

from datetime import datetime

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ._helpers import apply_header_row, calc_note, set_col_widths
from ._styles import (
    BOLD_FONT,
    BOTTOM_BORDER,
    DARK_BLUE,
    FORMULA_FONT,
    LIGHT_BLUE,
    LINK_FONT,
    NORMAL_FONT,
    PCT_FORMAT,
    PURPLE,
    THIN_BORDER,
    USD_M_FORMAT,
    YELLOW,
)

SN = "P&L & Cash Flow"
_MARGIN_FONT = Font(name="Calibri", size=9, italic=True, color="808080")


def _sum_formula(r, proj_years):
    return f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})"


def _write_currency_row(
    ws,
    r,
    label,
    formula_fn,
    proj_years,
    total_col,
    label_font=NORMAL_FONT,
    value_font=FORMULA_FONT,
    fill=None,
    bottom_border=False,
):
    """Write: col 1 label; cols 2..proj_years+1 formulas; total_col SUM."""
    ws.cell(row=r, column=1, value=label).font = label_font
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=formula_fn(y))
        c.font = value_font
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        if fill is not None:
            c.fill = fill
    c = ws.cell(row=r, column=total_col, value=_sum_formula(r, proj_years))
    c.font = FORMULA_FONT
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER
    if bottom_border:
        for col in range(1, total_col + 1):
            ws.cell(row=r, column=col).border = BOTTOM_BORDER


def _write_margin_row(ws, r, label, numer_row, denom_row, proj_years):
    """Italic grey ratio row: no SUM total, no border on col 1."""
    ws.cell(row=r, column=1, value=label).font = _MARGIN_FONT
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=IFERROR({cl}{numer_row}/{cl}{denom_row},0)")
        c.font = _MARGIN_FONT
        c.number_format = PCT_FORMAT
        c.border = THIN_BORDER


def _register_row(ws, r, tracker, key_prefix, proj_years):
    for y in range(proj_years):
        tracker.set(f"{key_prefix}.y{y}", SN, r, 2 + y)


def _write_header(ws, proj_years, base_year):
    r = 1
    ws.cell(row=r, column=1, value="PROFIT & LOSS — FORMULA-BASED ($M)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(ws, r, 1, "All lines are cross-sheet formulas. Change Assumptions to update.")
    r += 2

    headers = ["($M)"] + [str(base_year + y) for y in range(proj_years)] + ["Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    return r + 1


def _write_revenue_and_gp(ws, tracker, proj_years, total_col, r):
    # Revenue — cross-sheet link to Revenue Build
    _write_currency_row(
        ws,
        r,
        "Revenue",
        lambda y: f"={tracker.get(f'rev.total.y{y}')}",
        proj_years,
        total_col,
        label_font=BOLD_FONT,
        value_font=LINK_FONT,
    )
    _register_row(ws, r, tracker, "pl.rev", proj_years)
    rev_row = r
    r += 1

    # COGS — negated cross-sheet link
    _write_currency_row(
        ws,
        r,
        "(-) COGS",
        lambda y: f"=-{tracker.get(f'cost.cogs.y{y}')}",
        proj_years,
        total_col,
        value_font=LINK_FONT,
    )
    cogs_row = r
    r += 1

    # Gross Profit
    _write_currency_row(
        ws,
        r,
        "GROSS PROFIT",
        lambda y: f"={get_column_letter(2 + y)}{rev_row}+{get_column_letter(2 + y)}{cogs_row}",
        proj_years,
        total_col,
        label_font=BOLD_FONT,
        bottom_border=True,
    )
    _register_row(ws, r, tracker, "pl.gp", proj_years)
    gp_row = r
    r += 1

    _write_margin_row(ws, r, "  GP Margin %", gp_row, rev_row, proj_years)
    return r + 2, rev_row, gp_row


def _write_opex_and_ebit(ws, tracker, proj_years, total_col, rev_row, gp_row, r):
    blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    _write_currency_row(
        ws,
        r,
        "(-) R&D",
        lambda y: f"=-{tracker.get(f'cost.rd.y{y}')}",
        proj_years,
        total_col,
        value_font=LINK_FONT,
    )
    rd_row = r
    r += 1

    _write_currency_row(
        ws,
        r,
        "(-) SG&A",
        lambda y: f"=-{tracker.get(f'cost.sga.y{y}')}",
        proj_years,
        total_col,
        value_font=LINK_FONT,
    )
    sga_row = r
    r += 1

    _write_currency_row(
        ws,
        r,
        "EBIT",
        lambda y: (
            f"={get_column_letter(2 + y)}{gp_row}"
            f"+{get_column_letter(2 + y)}{rd_row}"
            f"+{get_column_letter(2 + y)}{sga_row}"
        ),
        proj_years,
        total_col,
        label_font=BOLD_FONT,
        fill=blue_fill,
        bottom_border=True,
    )
    _register_row(ws, r, tracker, "pl.ebit", proj_years)
    ebit_row = r
    r += 1

    _write_margin_row(ws, r, "  EBIT Margin %", ebit_row, rev_row, proj_years)
    return r + 2, ebit_row


def _write_tax_and_fcf(ws, tracker, proj_years, total_col, ebit_row, r):
    tax_ref = tracker.get("tax_rate")

    _write_currency_row(
        ws,
        r,
        "(-) Tax",
        lambda y: (
            f"=-IF({get_column_letter(2 + y)}{ebit_row}>0,"
            f"{get_column_letter(2 + y)}{ebit_row}*{tax_ref},0)"
        ),
        proj_years,
        total_col,
    )
    tax_row = r
    r += 1

    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    _write_currency_row(
        ws,
        r,
        "FREE CASH FLOW (Unrisked)",
        lambda y: f"={get_column_letter(2 + y)}{ebit_row}+{get_column_letter(2 + y)}{tax_row}",
        proj_years,
        total_col,
        label_font=BOLD_FONT,
        fill=yellow_fill,
    )
    _register_row(ws, r, tracker, "pl.fcf", proj_years)
    fcf_row = r
    return r + 2, fcf_row


def _write_cumulative_cf(ws, fcf_row, proj_years, r):
    ws.cell(row=r, column=1, value="Cumulative Cash Flow").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        if y == 0:
            formula = f"={cl}{fcf_row}"
        else:
            prev = get_column_letter(col - 1)
            formula = f"={prev}{r}+{cl}{fcf_row}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER


def build_pl_sheet(wb, config, tracker):
    ws = wb.create_sheet(SN)
    ws.sheet_properties.tabColor = PURPLE

    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)
    total_col = proj_years + 2

    set_col_widths(ws, {"A": 40})

    r = _write_header(ws, proj_years, base_year)
    r, rev_row, gp_row = _write_revenue_and_gp(ws, tracker, proj_years, total_col, r)
    r, ebit_row = _write_opex_and_ebit(ws, tracker, proj_years, total_col, rev_row, gp_row, r)
    r, fcf_row = _write_tax_and_fcf(ws, tracker, proj_years, total_col, ebit_row, r)
    _write_cumulative_cf(ws, fcf_row, proj_years, r)

    ws.freeze_panes = "B5"
    return ws
