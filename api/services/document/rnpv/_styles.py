"""Shared styling constants for the rNPV Excel model sheets."""

from openpyxl.styles import Border, Font, PatternFill, Side

DARK_BLUE = "1F4E79"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
INPUT_BLUE = "BDD7EE"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN = "548235"
LIGHT_GREEN = "E2EFDA"
RED = "C00000"
LIGHT_RED = "FCE4EC"
ORANGE = "ED7D31"
DARK_GRAY = "404040"
YELLOW = "FFF2CC"
PURPLE = "7030A0"
TEAL = "00B0F0"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
INPUT_FILL = PatternFill(start_color=INPUT_BLUE, end_color=INPUT_BLUE, fill_type="solid")
RESEARCH_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
NORMAL_FONT = Font(name="Calibri", size=10, color=DARK_GRAY)
BOLD_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
FORMULA_FONT = Font(name="Calibri", size=10, color="000000")
LINK_FONT = Font(name="Calibri", size=10, color="008000")
INPUT_FONT = Font(name="Calibri", size=10, color="0000FF")

PCT_FORMAT = "0.0%"
PCT2_FORMAT = "0.00%"
USD_FORMAT = "$#,##0"
USD_M_FORMAT = "$#,##0.0"
NUM_FORMAT = "#,##0"

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=DARK_BLUE))
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
