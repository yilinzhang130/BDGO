"""Sheet 4: Cost Structure — R&D, COGS, SG&A roll-up (formulas only)."""

from datetime import datetime

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ._helpers import apply_header_row, calc_note, section_title, set_col_widths
from ._styles import (
    BOLD_FONT,
    DARK_BLUE,
    FORMULA_FONT,
    LIGHT_BLUE,
    LIGHT_RED,
    NORMAL_FONT,
    ORANGE,
    THIN_BORDER,
    USD_M_FORMAT,
)

SN = "Cost Structure"


def _write_year_row(
    ws, r, label, formula_fn, proj_years, total_col, label_font=NORMAL_FONT, fill=None
):
    """Write one cost row: label in col 1, per-year formulas, SUM in total_col.

    formula_fn(y) returns the Excel formula string for column (2 + y).
    """
    ws.cell(row=r, column=1, value=label).font = label_font
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=formula_fn(y))
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        if fill is not None:
            c.fill = fill
    start = get_column_letter(2)
    end = get_column_letter(proj_years + 1)
    c = ws.cell(row=r, column=total_col, value=f"=SUM({start}{r}:{end}{r})")
    c.font = FORMULA_FONT
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER


def _register_row(ws, r, tracker, key_prefix, proj_years):
    """After a row is written at row=r, register each year's cell in the tracker."""
    for y in range(proj_years):
        tracker.set(f"{key_prefix}.y{y}", SN, r, 2 + y)


def _write_header(ws, config, proj_years, base_year):
    r = 1
    ws.cell(row=r, column=1, value="COST STRUCTURE — FORMULA-BASED ($M)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(
        ws,
        r,
        1,
        "All costs use formulas referencing Assumptions. R&D spread linearly over duration.",
    )
    r += 2

    headers = ["Cost Item"] + [str(base_year + y) for y in range(proj_years)] + ["Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    return r + 1


def _write_rd_section(ws, config, tracker, proj_years, total_col, r):
    costs = config["costs"]
    section_title(ws, r, 1, "R&D COSTS")
    r += 1

    rd_rows = []
    phase_count = config.get("_rd_phase_count", len(costs.get("rd_by_phase", [])))
    for p_idx in range(phase_count):
        phase = costs.get("rd_by_phase", [])[p_idx]
        phase_name = phase.get("phase", f"Phase {p_idx}")
        cost_ref = tracker.get(f"rd{p_idx}.cost")
        dur_ref = tracker.get(f"rd{p_idx}.duration")
        start_ref = tracker.get(f"rd{p_idx}.start")

        def phase_formula(y, c=cost_ref, d=dur_ref, s=start_ref):
            return f"=IF(AND({y}>={s},{y}<{s}+CEILING({d},1)),{c}/{d},0)"

        _write_year_row(ws, r, f"  {phase_name}", phase_formula, proj_years, total_col)
        rd_rows.append(r)
        r += 1

    if tracker.refs.get("cmc_total"):
        cmc_ref = tracker.get("cmc_total")
        _write_year_row(
            ws,
            r,
            "  CMC / Manufacturing",
            lambda y: f"=IF({y}<5,{cmc_ref}/5,0)",
            proj_years,
            total_col,
        )
        rd_rows.append(r)
        r += 1

    def rd_total_formula(y, rows=rd_rows):
        col = get_column_letter(2 + y)
        return "=" + "+".join(f"{col}{row}" for row in rows)

    fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    _write_year_row(
        ws,
        r,
        "  TOTAL R&D",
        rd_total_formula,
        proj_years,
        total_col,
        label_font=BOLD_FONT,
        fill=fill,
    )
    _register_row(ws, r, tracker, "cost.rd", proj_years)
    rd_total_row = r
    return r + 2, rd_total_row


def _write_cogs_section(ws, tracker, proj_years, total_col, r):
    section_title(ws, r, 1, "COST OF GOODS SOLD (COGS)")
    r += 1
    cogs_ref = tracker.get("cogs_margin")

    def cogs_formula(y):
        return f"={tracker.get(f'rev.total.y{y}')}*{cogs_ref}"

    _write_year_row(
        ws, r, "  Revenue x COGS%", cogs_formula, proj_years, total_col, label_font=BOLD_FONT
    )
    _register_row(ws, r, tracker, "cost.cogs", proj_years)
    return r + 2, r


def _sales_formula(y, first_launch, reps, cpr, r0, r1, r2):
    yfl = y - first_launch
    if yfl < -1:
        return "=0"
    if yfl == -1:
        return f"={reps}*{cpr}/1000*{r0}"
    if yfl == 0:
        return f"={reps}*{cpr}/1000*{r1}"
    return f"={reps}*{cpr}/1000*{r2}"


def _msl_formula(y, first_launch, ct, cost):
    yfl = y - first_launch
    if yfl < -2:
        return "=0"
    if yfl < 0:
        return f"={ct}*{cost}/1000*0.5"
    return f"={ct}*{cost}/1000"


def _marketing_formula(y, first_launch, congress, pubs, digital, prelaunch):
    yfl = y - first_launch
    if yfl in (-2, -1):
        return f"={prelaunch}/2"
    if yfl >= 0:
        return f"={congress}+{pubs}+{digital}"
    return "=0"


def _ga_formula(y, first_launch, ga_ref, rev_ref):
    yfl = y - first_launch
    if -3 <= yfl <= 0:
        return f"=MAX({rev_ref}*{ga_ref},2)"
    return f"={rev_ref}*{ga_ref}"


def _write_sga_components(ws, config, tracker, first_launch, proj_years, total_col, r):
    component_rows = []

    sga_cfg = config["costs"].get("sga", {})
    if sga_cfg.get("sales_team") and tracker.refs.get("sga.reps"):
        reps = tracker.get("sga.reps")
        cpr = tracker.get("sga.cost_per_rep")
        r0, r1, r2 = (tracker.get(f"sga.ramp{i}") for i in range(3))
        _write_year_row(
            ws,
            r,
            "  Sales Team",
            lambda y: _sales_formula(y, first_launch, reps, cpr, r0, r1, r2),
            proj_years,
            total_col,
        )
        component_rows.append(r)
        r += 1

    if tracker.refs.get("sga.msl_count"):
        ct = tracker.get("sga.msl_count")
        cost = tracker.get("sga.msl_cost")
        _write_year_row(
            ws,
            r,
            "  MSLs",
            lambda y: _msl_formula(y, first_launch, ct, cost),
            proj_years,
            total_col,
        )
        component_rows.append(r)
        r += 1

    if tracker.refs.get("sga.congress"):
        congress = tracker.get("sga.congress")
        pubs = tracker.get("sga.pubs")
        digital = tracker.get("sga.digital")
        prelaunch = tracker.get("sga.prelaunch")
        _write_year_row(
            ws,
            r,
            "  Marketing & Promotion",
            lambda y: _marketing_formula(y, first_launch, congress, pubs, digital, prelaunch),
            proj_years,
            total_col,
        )
        component_rows.append(r)
        r += 1

    if tracker.refs.get("sga.ga_pct"):
        ga_ref = tracker.get("sga.ga_pct")
        _write_year_row(
            ws,
            r,
            "  G&A",
            lambda y: _ga_formula(y, first_launch, ga_ref, tracker.get(f"rev.total.y{y}")),
            proj_years,
            total_col,
        )
        component_rows.append(r)
        r += 1

    return r, component_rows


def _write_sga_section(ws, config, tracker, proj_years, total_col, r):
    section_title(ws, r, 1, "SG&A")
    r += 1
    first_launch = min(ind.get("years_to_launch", 5) for ind in config["indications"])

    r, component_rows = _write_sga_components(
        ws, config, tracker, first_launch, proj_years, total_col, r
    )

    def sga_total_formula(y, rows=component_rows):
        if not rows:
            return "=0"
        col = get_column_letter(2 + y)
        return "=" + "+".join(f"{col}{row}" for row in rows)

    fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    _write_year_row(
        ws,
        r,
        "  TOTAL SG&A",
        sga_total_formula,
        proj_years,
        total_col,
        label_font=BOLD_FONT,
        fill=fill,
    )
    _register_row(ws, r, tracker, "cost.sga", proj_years)
    return r + 2, r


def _write_grand_total(ws, tracker, rd_row, cogs_row, sga_row, proj_years, total_col, r):
    def total_formula(y):
        col = get_column_letter(2 + y)
        return f"={col}{rd_row}+{col}{cogs_row}+{col}{sga_row}"

    fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
    _write_year_row(
        ws,
        r,
        "TOTAL COSTS ($M)",
        total_formula,
        proj_years,
        total_col,
        label_font=BOLD_FONT,
        fill=fill,
    )
    _register_row(ws, r, tracker, "cost.total", proj_years)


def build_cost_sheet(wb, config, tracker):
    ws = wb.create_sheet(SN)
    ws.sheet_properties.tabColor = ORANGE

    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)
    total_col = proj_years + 2

    set_col_widths(ws, {"A": 40})

    r = _write_header(ws, config, proj_years, base_year)
    r, rd_total_row = _write_rd_section(ws, config, tracker, proj_years, total_col, r)
    r, cogs_total_row = _write_cogs_section(ws, tracker, proj_years, total_col, r)
    r, sga_total_row = _write_sga_section(ws, config, tracker, proj_years, total_col, r)
    _write_grand_total(
        ws, tracker, rd_total_row, cogs_total_row, sga_total_row, proj_years, total_col, r
    )

    ws.freeze_panes = "B4"
    return ws
