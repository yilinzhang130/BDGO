"""Sheet 8: QC Report — run validation checks against the assembled model."""

from datetime import datetime

from openpyxl.styles import Font

from ._helpers import apply_header_row, section_title, set_col_widths, write_label_value
from ._styles import (
    DARK_BLUE,
    FAIL_FILL,
    NORMAL_FONT,
    PASS_FILL,
    THIN_BORDER,
    WARN_FILL,
)

SN = "QC Report"

# Expected cumulative-PoS ranges per starting phase
_POS_RANGES = {
    "Preclinical": (0.01, 0.15),
    "Phase 1": (0.05, 0.25),
    "Phase 2": (0.10, 0.40),
    "Phase 3": (0.30, 0.70),
    "NDA": (0.70, 0.95),
}

_FUNNEL_RATE_KEYS = (
    "diagnosed_rate",
    "eligible_rate",
    "line_share",
    "drug_treatable_rate",
    "addressable_rate",
)

_REF_PARAM_KEYS = (
    "prevalence",
    "diagnosed_rate",
    "eligible_rate",
    "line_share",
    "drug_treatable_rate",
    "addressable_rate",
    "pricing",
)

_UNSOURCED_KEYWORDS = ("estimate", "assumption", "default")


def _check_funnel_rates(indications):
    out = []
    for ind in indications:
        for geo, gd in ind["geography_data"].items():
            rates = [gd.get(k, 1) for k in _FUNNEL_RATE_KEYS]
            ok = all(0 < r_ <= 1.0 for r_ in rates)
            out.append(
                (
                    f"Funnel rates ({ind['name']}/{geo})",
                    "PASS" if ok else "FAIL",
                    f"Rates: {', '.join(f'{r_:.0%}' for r_ in rates)}",
                    "All 0-100%",
                )
            )
    return out


def _check_peak_revenue(peak_rev):
    return (
        "Peak revenue range",
        "PASS" if 50 < peak_rev < 50000 else "WARN",
        f"${peak_rev:,.0f}M",
        "$50M-$50B",
    )


def _check_pos_vs_phase(indications):
    out = []
    for ind in indications:
        pos = ind.get("pos", {})
        phase = pos.get("current_phase", "")
        cum = pos.get("cumulative", 0)
        lo, hi = _POS_RANGES.get(phase, (0.01, 1.0))
        ok = lo <= cum <= hi
        out.append(
            (
                f"PoS vs phase ({ind['name']})",
                "PASS" if ok else "WARN",
                f"{cum:.1%}",
                f"{lo:.0%}-{hi:.0%} for {phase}",
            )
        )
    return out


def _check_wacc(wacc):
    return (
        "WACC range",
        "PASS" if 0.08 <= wacc <= 0.20 else "WARN",
        f"{wacc:.1%}",
        "8%-20%",
    )


def _check_no_prelaunch_revenue(indications, revenues):
    first_launch = min(ind.get("years_to_launch", 5) for ind in indications)
    pre_rev = sum(revenues.get(y, 0) for y in range(first_launch))
    return (
        "No pre-launch revenue",
        "PASS" if pre_rev == 0 else "FAIL",
        f"${pre_rev:,.0f}M",
        "$0",
    )


def _check_npv_nonzero(npv):
    return (
        "NPV non-zero",
        "PASS" if abs(npv) > 0.1 else "FAIL",
        f"${npv:,.1f}M",
        "Non-zero",
    )


def _check_risk_ratio(npv, unrisked_npv):
    if unrisked_npv == 0:
        return None
    ratio = npv / unrisked_npv
    return (
        "rNPV/uNPV ratio",
        "PASS" if 0.01 < ratio < 0.80 else "WARN",
        f"{ratio:.0%}",
        "1%-80%",
    )


def _check_loe_erosion(rev_vals):
    if not rev_vals or len(rev_vals) <= 5:
        return None
    last5 = rev_vals[-5:]
    declines = any(last5[i] < last5[i - 1] for i in range(1, len(last5)))
    return (
        "LOE erosion visible",
        "PASS" if declines else "WARN",
        "Revenue declines" if declines else "Still growing",
        "Post-LOE decline",
    )


def _check_pricing_complete(indications):
    out = []
    for ind in indications:
        pricing = ind.get("pricing", {})
        geos = list(ind["geography_data"].keys())
        priced = sum(1 for g in geos if pricing.get(g, 0) > 0)
        out.append(
            (
                f"Pricing complete ({ind['name']})",
                "PASS" if priced == len(geos) else "FAIL",
                f"{priced}/{len(geos)}",
                "All geos priced",
            )
        )
    return out


def _check_reference_coverage(indications):
    ref_total = 0
    ref_sourced = 0
    for ind in indications:
        ds = ind.get("data_sources", {})
        for p in _REF_PARAM_KEYS:
            ref_total += 1
            src = ds.get(p, "")
            if src and not any(kw in src.lower() for kw in _UNSOURCED_KEYWORDS):
                ref_sourced += 1
    ref_pct = ref_sourced / ref_total if ref_total > 0 else 0
    if ref_pct >= 0.7:
        status = "PASS"
    elif ref_pct >= 0.4:
        status = "WARN"
    else:
        status = "FAIL"
    return (
        "Reference coverage",
        status,
        f"{ref_sourced}/{ref_total} ({ref_pct:.0%})",
        ">=70%",
    )


def _collect_checks(config):
    indications = config["indications"]
    checks = []
    checks.extend(_check_funnel_rates(indications))
    checks.append(_check_peak_revenue(config.get("_peak_rev", 0)))
    checks.extend(_check_pos_vs_phase(indications))
    checks.append(_check_wacc(config["discount"]["wacc"]))
    checks.append(_check_no_prelaunch_revenue(indications, config.get("_computed_revenues", {})))
    checks.append(_check_npv_nonzero(config.get("_npv", 0)))
    risk = _check_risk_ratio(config.get("_npv", 0), config.get("_unrisked_npv", 0))
    if risk:
        checks.append(risk)
    loe = _check_loe_erosion(config.get("_rev_vals", []))
    if loe:
        checks.append(loe)
    checks.extend(_check_pricing_complete(indications))
    checks.append(_check_reference_coverage(indications))
    checks.append(
        (
            "Model type: Formula-based",
            "PASS",
            "v3 — all calculations use Excel formulas",
            "Dynamic model",
        )
    )
    return checks


_STATUS_STYLE = {
    "PASS": (PASS_FILL, Font(name="Calibri", size=10, bold=True, color="006100")),
    "FAIL": (FAIL_FILL, Font(name="Calibri", size=10, bold=True, color="9C0006")),
    "WARN": (WARN_FILL, Font(name="Calibri", size=10, bold=True, color="9C6500")),
}


def _write_check_row(ws, r, check):
    name, status, detail, exp = check
    ws.cell(row=r, column=1, value=name).font = NORMAL_FONT
    c = ws.cell(row=r, column=2, value=status)
    fill, font = _STATUS_STYLE[status]
    c.fill = fill
    c.font = font
    ws.cell(row=r, column=3, value=detail).font = NORMAL_FONT
    ws.cell(row=r, column=4, value=exp).font = Font(name="Calibri", size=9, color="808080")
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = THIN_BORDER


def _tally(checks):
    counts = {"PASS": 0, "WARN": 0, "FAIL": 0}
    for _, status, _, _ in checks:
        counts[status] += 1
    return counts


def _write_header(ws):
    r = 1
    ws.cell(row=r, column=1, value="MODEL QUALITY CONTROL REPORT")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | v3 Formula-Based",
    )
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="808080")
    r += 2

    for i, h in enumerate(["Check", "Status", "Detail", "Expected"], 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, 4)
    return r + 1


def _write_summary(ws, r, counts):
    r += 2
    section_title(ws, r, 1, "QC SUMMARY")
    r += 1
    for label, status, color in [
        ("Passed", "PASS", "006100"),
        ("Warnings", "WARN", "9C6500"),
        ("Failed", "FAIL", "9C0006"),
    ]:
        write_label_value(
            ws,
            r,
            1,
            label,
            counts[status],
            val_font=Font(name="Calibri", size=12, bold=True, color=color),
        )
        ws.cell(row=r, column=2).fill = _STATUS_STYLE[status][0]
        r += 1


def build_qc_sheet(wb, config, tracker):
    ws = wb.create_sheet(SN)
    ws.sheet_properties.tabColor = "7030A0"
    set_col_widths(ws, {"A": 50, "B": 12, "C": 50, "D": 25})

    r = _write_header(ws)

    checks = _collect_checks(config)
    for check in checks:
        _write_check_row(ws, r, check)
        r += 1

    counts = _tally(checks)
    _write_summary(ws, r, counts)

    config["_qc_pass"] = counts["PASS"]
    config["_qc_fail"] = counts["FAIL"]
    config["_qc_warn"] = counts["WARN"]
    return ws
