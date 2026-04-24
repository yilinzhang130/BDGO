"""Sheet 3: Revenue Build — treated patients × net price for each indication+geography."""

from datetime import datetime

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ._helpers import apply_header_row, calc_note, section_title, set_col_widths
from ._styles import (
    BOLD_FONT,
    BOTTOM_BORDER,
    DARK_BLUE,
    FORMULA_FONT,
    GREEN,
    LIGHT_GREEN,
    LINK_FONT,
    MED_BLUE,
    NORMAL_FONT,
    NUM_FORMAT,
    THIN_BORDER,
    USD_FORMAT,
    USD_M_FORMAT,
    YELLOW,
)

SN = "Revenue Build"


def _sum_formula(r, proj_years):
    return f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})"


def _write_year_row(
    ws,
    r,
    label,
    formula_fn,
    proj_years,
    total_col=None,
    label_font=NORMAL_FONT,
    value_font=LINK_FONT,
    value_fmt=USD_M_FORMAT,
    fill=None,
    bottom_border=False,
):
    """Label in col 1, per-year formulas, optional SUM in total_col, optional fill/border."""
    ws.cell(row=r, column=1, value=label).font = label_font
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=formula_fn(y))
        c.font = value_font
        c.number_format = value_fmt
        c.border = THIN_BORDER
        if fill is not None:
            c.fill = fill
    if total_col is not None:
        c = ws.cell(row=r, column=total_col, value=_sum_formula(r, proj_years))
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
    if bottom_border:
        end = total_col or (proj_years + 1)
        for col in range(1, end + 1):
            ws.cell(row=r, column=col).border = BOTTOM_BORDER


def _write_header(ws):
    r = 1
    ws.cell(row=r, column=1, value="REVENUE BUILD — FORMULA-BASED ($M)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(ws, r, 1, "Revenue = Treated Patients x Net Price / 1,000,000. All via formulas.")
    return r + 2


def _write_indication_header(ws, ind, r, proj_years, base_year):
    title = f"INDICATION: {ind['name']}"
    if ind.get("line_of_therapy"):
        title += f"  ({ind['line_of_therapy']})"
    section_title(ws, r, 1, title)
    r += 1

    headers = [""] + [str(base_year + y) for y in range(proj_years)] + ["Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    return r + 1


def _write_geo_block(ws, ind_idx, geo, tracker, proj_years, total_col, r):
    """Treated / x Net Price / -> Revenue triplet for one geography. Returns (r, rev_row)."""
    ws.cell(row=r, column=1, value=f"  {geo}").font = Font(
        name="Calibri", size=10, bold=True, color=MED_BLUE
    )
    r += 1

    # Treated Patients — cross-sheet link to Patient Funnel
    _write_year_row(
        ws,
        r,
        "    Treated Patients",
        lambda y: f"={tracker.get(f'funnel.ind{ind_idx}.{geo}.treated.y{y}')}",
        proj_years,
        value_fmt=NUM_FORMAT,
    )
    treated_row = r
    r += 1

    # x Net Price — cross-sheet link to Assumptions (same net price every year)
    net_price_ref = tracker.get(f"ind{ind_idx}.{geo}.net_price")
    _write_year_row(
        ws,
        r,
        "    x Net Price ($)",
        lambda _y: f"={net_price_ref}",
        proj_years,
        value_fmt=USD_FORMAT,
    )
    price_row = r
    r += 1

    # Revenue ($M) = Treated x Net Price / 1e6
    def rev_formula(y, tr=treated_row, pr=price_row):
        cl = get_column_letter(2 + y)
        return f"={cl}{tr}*{cl}{pr}/1000000"

    _write_year_row(
        ws,
        r,
        "    -> Revenue ($M)",
        rev_formula,
        proj_years,
        total_col=total_col,
        label_font=BOLD_FONT,
        value_font=FORMULA_FONT,
        bottom_border=True,
    )
    for y in range(proj_years):
        tracker.set(f"rev.ind{ind_idx}.{geo}.y{y}", SN, r, 2 + y)
    return r + 2, r


def _write_indication_total(ws, r, ind_idx, ind_name, geo_rev_rows, tracker, proj_years, total_col):
    def total_formula(y, rows=geo_rev_rows):
        col = get_column_letter(2 + y)
        return "=" + "+".join(f"{col}{row}" for row in rows)

    green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    _write_year_row(
        ws,
        r,
        f"  TOTAL {ind_name} ($M)",
        total_formula,
        proj_years,
        total_col=total_col,
        label_font=BOLD_FONT,
        value_font=FORMULA_FONT,
        fill=green_fill,
        bottom_border=True,
    )
    for y in range(proj_years):
        tracker.set(f"rev.ind{ind_idx}.total.y{y}", SN, r, 2 + y)


def _write_indication_section(ws, ind_idx, ind, tracker, proj_years, base_year, total_col, r):
    r = _write_indication_header(ws, ind, r, proj_years, base_year)

    geo_rev_rows = []
    for geo in ind["geography_data"]:
        r, rev_row = _write_geo_block(ws, ind_idx, geo, tracker, proj_years, total_col, r)
        geo_rev_rows.append(rev_row)

    _write_indication_total(
        ws, r, ind_idx, ind["name"], geo_rev_rows, tracker, proj_years, total_col
    )
    return r + 2


def _write_grand_total(ws, indications, tracker, proj_years, base_year, total_col, r):
    section_title(ws, r, 1, "TOTAL REVENUE — ALL INDICATIONS ($M)")
    r += 1
    headers = [""] + [str(base_year + y) for y in range(proj_years)] + ["Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    def grand_formula(y):
        return "=" + "+".join(
            tracker.local(f"rev.ind{idx}.total.y{y}") for idx in range(len(indications))
        )

    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    _write_year_row(
        ws,
        r,
        "Total Revenue ($M)",
        grand_formula,
        proj_years,
        total_col=total_col,
        label_font=BOLD_FONT,
        value_font=FORMULA_FONT,
        fill=yellow_fill,
    )
    for y in range(proj_years):
        tracker.set(f"rev.total.y{y}", SN, r, 2 + y)
    return r + 2


def build_revenue_sheet(wb, config, tracker):
    ws = wb.create_sheet(SN)
    ws.sheet_properties.tabColor = GREEN
    set_col_widths(ws, {"A": 38})

    indications = config["indications"]
    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)
    total_col = proj_years + 2

    r = _write_header(ws)
    for ind_idx, ind in enumerate(indications):
        r = _write_indication_section(
            ws, ind_idx, ind, tracker, proj_years, base_year, total_col, r
        )
    _write_grand_total(ws, indications, tracker, proj_years, base_year, total_col, r)

    ws.freeze_panes = "B4"
    return ws
