"""Shared cell-writing helpers used by every rNPV sheet builder."""

import math

from openpyxl.styles import Alignment, Font

from ._styles import (
    BOLD_FONT,
    DARK_BLUE,
    FORMULA_FONT,
    HEADER_FILL,
    HEADER_FONT,
    INPUT_FILL,
    INPUT_FONT,
    NORMAL_FONT,
    SECTION_FONT,
    SUBHEADER_FILL,
    SUBHEADER_FONT,
    THIN_BORDER,
)


def s_curve(year, peak, ramp_years=7):
    if year <= 0:
        return 0
    if year >= ramp_years:
        return peak
    midpoint = ramp_years / 2
    steepness = 1.2
    raw = 1 / (1 + math.exp(-steepness * (year - midpoint)))
    raw_min = 1 / (1 + math.exp(-steepness * (1 - midpoint)))
    raw_max = 1 / (1 + math.exp(-steepness * (ramp_years - midpoint)))
    normalized = (raw - raw_min) / (raw_max - raw_min)
    return peak * normalized


def apply_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_subheader_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def write_row(ws, row, data, font=None, fill=None, num_fmt=None, bold_first=False):
    for i, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = font or NORMAL_FONT
        if fill:
            cell.fill = fill
        if num_fmt and i > 1:
            cell.number_format = num_fmt
        if bold_first and i == 1:
            cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right" if i > 1 else "left", vertical="center")


def write_formula_row(ws, row, label, formulas, font=None, fill=None, num_fmt=None):
    """Write a row where col 1 = label, col 2+ = Excel formulas."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = BOLD_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for i, f in enumerate(formulas, 2):
        cell = ws.cell(row=row, column=i, value=f)
        cell.font = font or FORMULA_FONT
        if fill:
            cell.fill = fill
        if num_fmt:
            cell.number_format = num_fmt
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center")


def write_label_value(
    ws, row, col, label, val, fmt=None, label_font=None, val_font=None, fill=None
):
    c1 = ws.cell(row=row, column=col, value=label)
    c1.font = label_font or BOLD_FONT
    c1.border = THIN_BORDER
    c2 = ws.cell(row=row, column=col + 1, value=val)
    c2.font = val_font or Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
    if fmt:
        c2.number_format = fmt
    if fill:
        c2.fill = fill
    c2.border = THIN_BORDER


def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def section_title(ws, row, col, text):
    ws.cell(row=row, column=col, value=text).font = SECTION_FONT


def calc_note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=9, italic=True, color="808080")


def write_input_cell(ws, row, col, val, fmt=None, tracker=None, key=None, sheet_name=None):
    """Write a blue-font input cell and optionally register in tracker."""
    c = ws.cell(row=row, column=col, value=val)
    c.font = INPUT_FONT
    c.fill = INPUT_FILL
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    if tracker and key and sheet_name:
        tracker.set(key, sheet_name, row, col)
    return c
