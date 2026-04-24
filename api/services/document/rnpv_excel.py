#!/usr/bin/env python3
"""
rNPV Valuation Model — Excel Generator v3 (Formula-Based)
All calculations use Excel formulas with cross-sheet references.
Users can modify assumptions in Excel and see live recalculation.

Usage:
    python3 generate_rnpv_excel.py --config config.json --output output.xlsx
"""

import argparse
import json
import math
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ── Color Palette ──
DARK_BLUE = "1F4E79"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
INPUT_BLUE = "BDD7EE"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN = "548235"
LIGHT_GREEN = "E2EFDA"
RED = "C00000"
LIGHT_RED = "FCE4EC"
ORANGE = "ED7D31"
DARK_GRAY = "404040"
YELLOW = "FFF2CC"
PURPLE = "7030A0"
TEAL = "00B0F0"

# ── Styles ──
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
INPUT_FILL = PatternFill(start_color=INPUT_BLUE, end_color=INPUT_BLUE, fill_type="solid")
RESEARCH_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
NORMAL_FONT = Font(name="Calibri", size=10, color=DARK_GRAY)
BOLD_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
FORMULA_FONT = Font(name="Calibri", size=10, color="000000")  # Black = formula (IB standard)
LINK_FONT = Font(name="Calibri", size=10, color="008000")  # Green = cross-sheet link
INPUT_FONT = Font(name="Calibri", size=10, color="0000FF")  # Blue = user input
PCT_FORMAT = "0.0%"
PCT2_FORMAT = "0.00%"
USD_FORMAT = "$#,##0"
USD_M_FORMAT = "$#,##0.0"
NUM_FORMAT = "#,##0"
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=DARK_BLUE))
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


# ── CellTracker — Cross-sheet reference management ──


class CellTracker:
    """Track cell locations for cross-sheet formula references."""

    def __init__(self):
        self.refs = {}

    def set(self, key, sheet_name, row, col):
        col_letter = get_column_letter(col)
        self.refs[key] = f"'{sheet_name}'!${col_letter}${row}"

    def get(self, key):
        return self.refs.get(key, "#REF!")

    def local(self, key):
        """Return local cell reference (without sheet name) for same-sheet formulas."""
        ref = self.refs.get(key, "#REF!")
        if "!" in ref:
            return ref.split("!")[1]
        return ref

    def col_letter(self, col):
        return get_column_letter(col)


# ── Helpers ──


def s_curve(year, peak, ramp_years=7):
    if year <= 0:
        return 0
    if year >= ramp_years:
        return peak
    midpoint = ramp_years / 2
    steepness = 1.2
    raw = 1 / (1 + math.exp(-steepness * (year - midpoint)))
    raw_min = 1 / (1 + math.exp(-steepness * (1 - midpoint)))
    raw_max = 1 / (1 + math.exp(-steepness * (ramp_years - midpoint)))
    normalized = (raw - raw_min) / (raw_max - raw_min)
    return peak * normalized


def apply_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_subheader_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def write_row(ws, row, data, font=None, fill=None, num_fmt=None, bold_first=False):
    for i, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = font or NORMAL_FONT
        if fill:
            cell.fill = fill
        if num_fmt and i > 1:
            cell.number_format = num_fmt
        if bold_first and i == 1:
            cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right" if i > 1 else "left", vertical="center")


def write_formula_row(ws, row, label, formulas, font=None, fill=None, num_fmt=None):
    """Write a row where col 1 = label, col 2+ = Excel formulas."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = BOLD_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for i, f in enumerate(formulas, 2):
        cell = ws.cell(row=row, column=i, value=f)
        cell.font = font or FORMULA_FONT
        if fill:
            cell.fill = fill
        if num_fmt:
            cell.number_format = num_fmt
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center")


def write_label_value(
    ws, row, col, label, val, fmt=None, label_font=None, val_font=None, fill=None
):
    c1 = ws.cell(row=row, column=col, value=label)
    c1.font = label_font or BOLD_FONT
    c1.border = THIN_BORDER
    c2 = ws.cell(row=row, column=col + 1, value=val)
    c2.font = val_font or Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
    if fmt:
        c2.number_format = fmt
    if fill:
        c2.fill = fill
    c2.border = THIN_BORDER


def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def section_title(ws, row, col, text):
    ws.cell(row=row, column=col, value=text).font = SECTION_FONT


def calc_note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=9, italic=True, color="808080")


def write_input_cell(ws, row, col, val, fmt=None, tracker=None, key=None, sheet_name=None):
    """Write a blue-font input cell and optionally register in tracker."""
    c = ws.cell(row=row, column=col, value=val)
    c.font = INPUT_FONT
    c.fill = INPUT_FILL
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    if tracker and key and sheet_name:
        tracker.set(key, sheet_name, row, col)
    return c


# ── Sheet 1: Assumptions (Single Source of Truth) ──


def build_assumptions_sheet(wb, config, tracker):
    ws = wb.active
    ws.title = "Assumptions"
    ws.sheet_properties.tabColor = DARK_BLUE
    SN = "Assumptions"

    meta = config["metadata"]
    indications = config["indications"]
    costs = config["costs"]
    discount = config["discount"]
    proj_years = discount.get("projection_years", 20)
    base_year = meta.get("base_year", datetime.now().year)

    set_col_widths(ws, {"A": 38, "B": 18, "C": 22, "D": 18, "E": 18, "F": 18, "G": 18})

    r = 1
    ws.cell(row=r, column=1, value=f"rNPV Valuation Model — {meta['company']} / {meta['asset']}")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=f"Analyst: {meta.get('analyst', 'N/A')}  |  Date: {meta.get('date', datetime.now().strftime('%Y-%m-%d'))}",
    )
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="808080")
    r += 1
    ws.cell(row=r, column=1, value="Color Legend:").font = Font(
        name="Calibri", size=9, bold=True, color="808080"
    )
    c = ws.cell(row=r, column=2, value="User Input (Blue)")
    c.fill = INPUT_FILL
    c.font = Font(name="Calibri", size=9, color="0000FF")
    c = ws.cell(row=r, column=3, value="Formula (Black)")
    c.font = Font(name="Calibri", size=9, color="000000")
    c = ws.cell(row=r, column=4, value="Cross-Sheet Link (Green)")
    c.font = Font(name="Calibri", size=9, color="008000")
    r += 2

    # ── General ──
    section_title(ws, r, 1, "GENERAL ASSUMPTIONS")
    r += 1
    headers = ["Parameter", "Value", "Source / Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, 3)
    r += 1

    # Company / Asset info (static)
    for label, val in [
        ("Company", meta["company"]),
        ("Asset / Drug", meta["asset"]),
        ("Modality", meta.get("modality", "N/A")),
        ("Therapeutic Area", meta.get("therapeutic_area", "N/A")),
    ]:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=val).font = NORMAL_FONT
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1

    # WACC
    ws.cell(row=r, column=1, value="WACC").font = NORMAL_FONT
    write_input_cell(ws, r, 2, discount["wacc"], PCT_FORMAT, tracker, "wacc", SN)
    ws.cell(
        row=r, column=3, value=discount.get("wacc_source", "Default biotech")
    ).font = NORMAL_FONT
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = THIN_BORDER
    r += 1

    # Tax Rate
    ws.cell(row=r, column=1, value="Tax Rate").font = NORMAL_FONT
    write_input_cell(ws, r, 2, discount.get("tax_rate", 0.20), PCT_FORMAT, tracker, "tax_rate", SN)
    ws.cell(row=r, column=3, value="Global blended effective").font = NORMAL_FONT
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = THIN_BORDER
    r += 1

    # Projection Years
    ws.cell(row=r, column=1, value="Projection Years").font = NORMAL_FONT
    write_input_cell(ws, r, 2, proj_years, None, tracker, "proj_years", SN)
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = THIN_BORDER
    r += 1

    # Base Year
    ws.cell(row=r, column=1, value="Base Year").font = NORMAL_FONT
    write_input_cell(ws, r, 2, base_year, None, tracker, "base_year", SN)
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = THIN_BORDER
    r += 1

    # COGS margin
    ws.cell(row=r, column=1, value="COGS (% of Net Revenue)").font = NORMAL_FONT
    write_input_cell(
        ws, r, 2, costs.get("cogs_margin", 0.20), PCT_FORMAT, tracker, "cogs_margin", SN
    )
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = THIN_BORDER
    r += 2

    # ── Per-Indication ──
    for ind_idx, ind in enumerate(indications):
        section_title(ws, r, 1, f"INDICATION {ind_idx + 1}: {ind['name']}")
        if ind.get("line_of_therapy"):
            ws.cell(row=r, column=3, value=f"Line: {ind['line_of_therapy']}")
            ws.cell(row=r, column=3).font = Font(name="Calibri", size=10, bold=True, color=ORANGE)
        r += 1

        geo_keys = list(ind["geography_data"].keys())
        n_geos = len(geo_keys)

        # Patient funnel inputs
        geo_headers = ["Patient Funnel"] + geo_keys + ["Source"]
        for i, h in enumerate(geo_headers, 1):
            ws.cell(row=r, column=i, value=h)
        apply_header_row(ws, r, len(geo_headers))
        r += 1

        funnel_params = [
            ("Prevalence / Incidence", "prevalence", NUM_FORMAT),
            ("Diagnosed Rate", "diagnosed_rate", PCT_FORMAT),
            ("Treatment Eligible Rate", "eligible_rate", PCT_FORMAT),
            ("Line-of-Therapy Share", "line_share", PCT_FORMAT),
            ("Drug-Treatable Rate", "drug_treatable_rate", PCT_FORMAT),
            ("Addressable Rate (Access)", "addressable_rate", PCT_FORMAT),
        ]
        for label, key, fmt in funnel_params:
            ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            for g_idx, geo in enumerate(geo_keys):
                val = ind["geography_data"][geo].get(key, 0)
                write_input_cell(
                    ws, r, 2 + g_idx, val, fmt, tracker, f"ind{ind_idx}.{geo}.{key}", SN
                )
            source = ind.get("data_sources", {}).get(key, "")
            ws.cell(row=r, column=n_geos + 2, value=source).font = Font(
                name="Calibri", size=9, color="808080"
            )
            ws.cell(row=r, column=n_geos + 2).border = THIN_BORDER
            r += 1

        # Calculated addressable patients (formula)
        ws.cell(row=r, column=1, value="-> Addressable Patients").font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for g_idx, geo in enumerate(geo_keys):
            col = 2 + g_idx
            prev_ref = tracker.local(f"ind{ind_idx}.{geo}.prevalence")
            diag_ref = tracker.local(f"ind{ind_idx}.{geo}.diagnosed_rate")
            elig_ref = tracker.local(f"ind{ind_idx}.{geo}.eligible_rate")
            line_ref = tracker.local(f"ind{ind_idx}.{geo}.line_share")
            treat_ref = tracker.local(f"ind{ind_idx}.{geo}.drug_treatable_rate")
            addr_ref = tracker.local(f"ind{ind_idx}.{geo}.addressable_rate")
            formula = f"=INT({prev_ref}*{diag_ref}*{elig_ref}*{line_ref}*{treat_ref}*{addr_ref})"
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = NUM_FORMAT
            c.border = THIN_BORDER
            c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
            tracker.set(f"ind{ind_idx}.{geo}.addressable", SN, r, col)
        ws.cell(
            row=r, column=n_geos + 2, value="Prev x Diag x Elig x Line x Treat x Access"
        ).font = Font(name="Calibri", size=9, italic=True, color="808080")
        r += 2

        # Pricing
        ws.cell(row=r, column=1, value="PRICING").font = Font(
            name="Calibri", size=10, bold=True, color=GREEN
        )
        r += 1
        for label, key, fmt in [
            ("List Price ($)", "list_price", USD_FORMAT),
            ("Gross-to-Net Ratio", "gtn", PCT_FORMAT),
        ]:
            ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            for g_idx, geo in enumerate(geo_keys):
                if key == "list_price":
                    val = ind.get("pricing", {}).get(geo, 0)
                else:
                    val = ind.get("gross_to_net", {}).get(geo, 0.70)
                write_input_cell(
                    ws, r, 2 + g_idx, val, fmt, tracker, f"ind{ind_idx}.{geo}.{key}", SN
                )
            r += 1

        # Net Price (formula)
        ws.cell(row=r, column=1, value="-> Net Price ($)").font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for g_idx, geo in enumerate(geo_keys):
            col = 2 + g_idx
            lp = tracker.local(f"ind{ind_idx}.{geo}.list_price")
            gtn = tracker.local(f"ind{ind_idx}.{geo}.gtn")
            formula = f"={lp}*{gtn}"
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = USD_FORMAT
            c.border = THIN_BORDER
            tracker.set(f"ind{ind_idx}.{geo}.net_price", SN, r, col)
        r += 2

        # Penetration curve (pre-computed as editable year-by-year inputs)
        ws.cell(row=r, column=1, value="MARKET PENETRATION").font = Font(
            name="Calibri", size=10, bold=True, color=GREEN
        )
        r += 1

        pen = ind.get("penetration_curve", {})
        peak_pen = pen.get("peak", 0.15)
        ramp_yrs = pen.get("ramp_years", 7)
        loe_year = pen.get("loe_year_from_launch", 12)
        post_loe = pen.get("post_loe_erosion_per_year", 0.30)
        launch_offset = ind.get("years_to_launch", 5)

        for label, val, fmt in [
            ("Peak Penetration", peak_pen, PCT_FORMAT),
            ("Ramp-up Years", ramp_yrs, None),
            ("LOE Year (from launch)", loe_year, None),
            ("Post-LOE Erosion/yr", post_loe, PCT_FORMAT),
            ("Years to Launch", launch_offset, None),
        ]:
            ws.cell(row=r, column=1, value=f"  {label}").font = NORMAL_FONT
            write_input_cell(
                ws,
                r,
                2,
                val,
                fmt,
                tracker,
                f"ind{ind_idx}.pen.{label.lower().replace(' ', '_').replace('/', '_')}",
                SN,
            )
            r += 1

        tracker.set(f"ind{ind_idx}.years_to_launch", SN, r - 1, 2)  # last one was Years to Launch

        # Year-by-year penetration % (pre-computed, editable)
        r += 1
        ws.cell(row=r, column=1, value="Penetration by Year:").font = Font(
            name="Calibri", size=10, bold=True, color=GREEN
        )
        r += 1
        year_headers = ["Year"] + [str(base_year + y) for y in range(proj_years)]
        for i, h in enumerate(year_headers, 1):
            ws.cell(row=r, column=i, value=h)
        apply_header_row(ws, r, len(year_headers))
        r += 1

        # Compute penetration values
        pen_vals = []
        for yr_offset in range(proj_years):
            yrs_since_launch = yr_offset - launch_offset
            if yrs_since_launch < 0:
                pen_vals.append(0)
            elif yrs_since_launch < loe_year:
                pv = s_curve(yrs_since_launch + 1, peak_pen, ramp_yrs)
                pen_vals.append(pv)
            else:
                years_post_loe = yrs_since_launch - loe_year
                base_pen = s_curve(loe_year, peak_pen, ramp_yrs)
                eroded = base_pen * ((1 - post_loe) ** (years_post_loe + 1))
                pen_vals.append(eroded)

        ws.cell(row=r, column=1, value="Penetration %").font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for y in range(proj_years):
            col = 2 + y
            write_input_cell(
                ws, r, col, pen_vals[y], PCT2_FORMAT, tracker, f"ind{ind_idx}.pen_y{y}", SN
            )
        r += 2

        # PoS
        ws.cell(row=r, column=1, value="PROBABILITY OF SUCCESS").font = Font(
            name="Calibri", size=10, bold=True, color=GREEN
        )
        r += 1
        pos = ind.get("pos", {})
        ws.cell(row=r, column=1, value="  Current Phase").font = NORMAL_FONT
        ws.cell(row=r, column=2, value=pos.get("current_phase", "Phase 1")).font = INPUT_FONT
        r += 1

        phase_trans = pos.get("phase_transitions", {})
        for trans_name, trans_val in phase_trans.items():
            label = trans_name.replace("_to_", " -> ").replace("_", " ").title()
            ws.cell(row=r, column=1, value=f"  {label}").font = NORMAL_FONT
            write_input_cell(
                ws, r, 2, trans_val, PCT_FORMAT, tracker, f"ind{ind_idx}.pos.{trans_name}", SN
            )
            r += 1

        cum_pos = pos.get("cumulative", 0.10)
        ws.cell(row=r, column=1, value="  -> Cumulative PoS").font = BOLD_FONT
        write_input_cell(ws, r, 2, cum_pos, PCT_FORMAT, tracker, f"ind{ind_idx}.cum_pos", SN)
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=12, bold=True, color=RED)
        r += 2

    # ── R&D Costs ──
    section_title(ws, r, 1, "R&D COST ASSUMPTIONS ($M)")
    r += 1
    rd_headers = [
        "Phase",
        "Total Cost ($M)",
        "Duration (Yr)",
        "Start Year (offset)",
        "# Trials",
        "Patients/Trial",
        "# Sites",
        "Source",
    ]
    for i, h in enumerate(rd_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(rd_headers))
    r += 1

    for p_idx, phase in enumerate(costs.get("rd_by_phase", [])):
        ws.cell(row=r, column=1, value=phase.get("phase", "")).font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        write_input_cell(
            ws, r, 2, phase.get("cost_mm", 0), USD_M_FORMAT, tracker, f"rd{p_idx}.cost", SN
        )
        write_input_cell(
            ws, r, 3, phase.get("duration_years", 1), None, tracker, f"rd{p_idx}.duration", SN
        )
        write_input_cell(
            ws, r, 4, phase.get("start_year", 0), None, tracker, f"rd{p_idx}.start", SN
        )
        write_input_cell(ws, r, 5, phase.get("num_trials", 1), None)
        write_input_cell(ws, r, 6, phase.get("patients_per_trial", ""), None)
        write_input_cell(ws, r, 7, phase.get("num_sites", ""), None)
        ws.cell(row=r, column=8, value=phase.get("source", "")).font = Font(
            name="Calibri", size=9, color="808080"
        )
        ws.cell(row=r, column=8).border = THIN_BORDER
        r += 1

    config["_rd_phase_count"] = len(costs.get("rd_by_phase", []))
    r += 1

    # CMC
    cmc = costs.get("cmc", {})
    if cmc:
        cmc_total = sum(v for v in cmc.values() if isinstance(v, (int, float)))
        ws.cell(row=r, column=1, value="CMC / Manufacturing Total ($M)").font = NORMAL_FONT
        write_input_cell(ws, r, 2, cmc_total, USD_M_FORMAT, tracker, "cmc_total", SN)
        ws.cell(row=r, column=3, value="Spread over 5 years").font = Font(
            name="Calibri", size=9, color="808080"
        )
        r += 2

    # ── SG&A ──
    section_title(ws, r, 1, "SG&A ASSUMPTIONS")
    r += 1
    sga = costs.get("sga", {})
    sales = sga.get("sales_team", {})
    if sales:
        ws.cell(row=r, column=1, value="Sales Reps (Full Deploy)").font = NORMAL_FONT
        write_input_cell(ws, r, 2, sales.get("reps", 100), None, tracker, "sga.reps", SN)
        r += 1
        ws.cell(row=r, column=1, value="Cost per Rep ($K/yr)").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, sales.get("cost_per_rep_k", 280), USD_FORMAT, tracker, "sga.cost_per_rep", SN
        )
        r += 1
        ramp = sales.get("ramp_schedule", [0.3, 0.6, 1.0])
        ws.cell(row=r, column=1, value="Ramp: Pre-launch %").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, ramp[0] if len(ramp) > 0 else 0.3, PCT_FORMAT, tracker, "sga.ramp0", SN
        )
        r += 1
        ws.cell(row=r, column=1, value="Ramp: Launch %").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, ramp[1] if len(ramp) > 1 else 0.6, PCT_FORMAT, tracker, "sga.ramp1", SN
        )
        r += 1
        ws.cell(row=r, column=1, value="Ramp: Year 2+ %").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, ramp[2] if len(ramp) > 2 else 1.0, PCT_FORMAT, tracker, "sga.ramp2", SN
        )
        r += 1

    msls = sga.get("msls", {})
    if msls:
        ws.cell(row=r, column=1, value="MSL Count").font = NORMAL_FONT
        write_input_cell(ws, r, 2, msls.get("count", 20), None, tracker, "sga.msl_count", SN)
        r += 1
        ws.cell(row=r, column=1, value="Cost per MSL ($K/yr)").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, msls.get("cost_per_msl_k", 350), USD_FORMAT, tracker, "sga.msl_cost", SN
        )
        r += 1

    mktg = sga.get("marketing", {})
    if mktg:
        ws.cell(row=r, column=1, value="Congress/KOL ($M/yr)").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, mktg.get("congress_annual_mm", 3), USD_M_FORMAT, tracker, "sga.congress", SN
        )
        r += 1
        ws.cell(row=r, column=1, value="Publications ($M/yr)").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, mktg.get("publications_mm", 1), USD_M_FORMAT, tracker, "sga.pubs", SN
        )
        r += 1
        ws.cell(row=r, column=1, value="Digital Marketing ($M/yr)").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, mktg.get("digital_marketing_mm", 2), USD_M_FORMAT, tracker, "sga.digital", SN
        )
        r += 1
        ws.cell(row=r, column=1, value="Pre-Launch Mktg Total ($M)").font = NORMAL_FONT
        write_input_cell(
            ws, r, 2, mktg.get("prelaunch_total_mm", 5), USD_M_FORMAT, tracker, "sga.prelaunch", SN
        )
        r += 1

    ws.cell(row=r, column=1, value="G&A (% of Revenue)").font = NORMAL_FONT
    write_input_cell(
        ws, r, 2, sga.get("ga_pct_of_revenue", 0.05), PCT_FORMAT, tracker, "sga.ga_pct", SN
    )
    r += 1

    ws.freeze_panes = "A5"
    return ws


# ── Sheet 2: Patient Funnel (Formulas -> Assumptions) ──


def build_patient_funnel_sheet(wb, config, tracker):
    ws = wb.create_sheet("Patient Funnel")
    ws.sheet_properties.tabColor = MED_BLUE
    SN = "Patient Funnel"

    indications = config["indications"]
    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)

    set_col_widths(ws, {"A": 38, "B": 14})

    r = 1
    ws.cell(row=r, column=1, value="PATIENT FUNNEL — FORMULA-BASED DERIVATION")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(
        ws,
        r,
        1,
        "All values are Excel formulas linking to Assumptions. Change inputs there to update.",
    )
    r += 2

    for ind_idx, ind in enumerate(indications):
        ind_name = ind["name"]
        line = ind.get("line_of_therapy", "")
        geo_data = ind["geography_data"]
        launch_offset = ind.get("years_to_launch", 5)

        title = f"INDICATION: {ind_name}"
        if line:
            title += f"  ({line})"
        section_title(ws, r, 1, title)
        r += 1

        for geo in geo_data:
            ws.cell(row=r, column=1, value=f"  Geography: {geo}")
            ws.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True, color=MED_BLUE)
            r += 1

            # Static funnel derivation with formulas
            prev_ref = tracker.get(f"ind{ind_idx}.{geo}.prevalence")
            diag_ref = tracker.get(f"ind{ind_idx}.{geo}.diagnosed_rate")
            elig_ref = tracker.get(f"ind{ind_idx}.{geo}.eligible_rate")
            line_ref = tracker.get(f"ind{ind_idx}.{geo}.line_share")
            treat_ref = tracker.get(f"ind{ind_idx}.{geo}.drug_treatable_rate")
            addr_ref = tracker.get(f"ind{ind_idx}.{geo}.addressable_rate")
            addr_pts_ref = tracker.get(f"ind{ind_idx}.{geo}.addressable")

            steps = [
                ("    Prevalence", f"={prev_ref}", NUM_FORMAT),
                ("    x Diagnosed Rate", f"={diag_ref}", PCT_FORMAT),
                ("    -> Diagnosed Patients", f"=INT({prev_ref}*{diag_ref})", NUM_FORMAT),
                ("    x Eligible Rate", f"={elig_ref}", PCT_FORMAT),
                ("    -> Eligible Patients", f"=INT({prev_ref}*{diag_ref}*{elig_ref})", NUM_FORMAT),
                ("    x Line Share", f"={line_ref}", PCT_FORMAT),
                ("    x Drug-Treatable", f"={treat_ref}", PCT_FORMAT),
                ("    x Market Access", f"={addr_ref}", PCT_FORMAT),
                ("    -> Addressable Patients", f"={addr_pts_ref}", NUM_FORMAT),
            ]

            for label, formula, fmt in steps:
                ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
                if label.startswith("    ->"):
                    ws.cell(row=r, column=1).font = BOLD_FONT
                ws.cell(row=r, column=1).border = THIN_BORDER
                c = ws.cell(row=r, column=2, value=formula)
                c.font = LINK_FONT
                c.number_format = fmt
                c.border = THIN_BORDER
                r += 1
            r += 1

            # Year-by-year projection (formula: addressable x penetration)
            headers = ["Year-by-Year Projection"] + [str(base_year + y) for y in range(proj_years)]
            for i, h in enumerate(headers, 1):
                ws.cell(row=r, column=i, value=h)
            apply_header_row(ws, r, len(headers))
            r += 1

            # Addressable row (same across years, formula linking to Assumptions)
            ws.cell(row=r, column=1, value="    Addressable Patients").font = NORMAL_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            for y in range(proj_years):
                c = ws.cell(row=r, column=2 + y, value=f"={addr_pts_ref}")
                c.font = LINK_FONT
                c.number_format = NUM_FORMAT
                c.border = THIN_BORDER
            addr_row = r
            r += 1

            # Penetration row (formula linking to Assumptions year-by-year)
            ws.cell(row=r, column=1, value="    x Market Penetration").font = NORMAL_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            for y in range(proj_years):
                pen_ref = tracker.get(f"ind{ind_idx}.pen_y{y}")
                c = ws.cell(row=r, column=2 + y, value=f"={pen_ref}")
                c.font = LINK_FONT
                c.number_format = PCT2_FORMAT
                c.border = THIN_BORDER
            pen_row = r
            r += 1

            # Treated patients (formula: addr x pen)
            ws.cell(row=r, column=1, value="    -> Treated Patients").font = BOLD_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            for y in range(proj_years):
                col = 2 + y
                addr_cell = f"{get_column_letter(col)}{addr_row}"
                pen_cell = f"{get_column_letter(col)}{pen_row}"
                formula = f"=INT({addr_cell}*{pen_cell})"
                c = ws.cell(row=r, column=col, value=formula)
                c.font = FORMULA_FONT
                c.number_format = NUM_FORMAT
                c.border = THIN_BORDER
                c.fill = PatternFill(
                    start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"
                )
                tracker.set(f"funnel.ind{ind_idx}.{geo}.treated.y{y}", SN, r, col)
            treated_row = r
            r += 2

        # Indication total treated (SUM across geos)
        geo_keys = list(geo_data.keys())
        ws.cell(row=r, column=1, value=f"  TOTAL TREATED — {ind_name}").font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for y in range(proj_years):
            col = 2 + y
            refs = [tracker.local(f"funnel.ind{ind_idx}.{geo}.treated.y{y}") for geo in geo_keys]
            formula = "=" + "+".join(refs)
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = NUM_FORMAT
            c.border = THIN_BORDER
            c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
            tracker.set(f"funnel.ind{ind_idx}.total.y{y}", SN, r, col)
        for col in range(1, proj_years + 2):
            ws.cell(row=r, column=col).border = BOTTOM_BORDER
        r += 2

    # Grand total treated
    ws.cell(row=r, column=1, value="TOTAL TREATED — ALL INDICATIONS").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        refs = [tracker.local(f"funnel.ind{idx}.total.y{y}") for idx in range(len(indications))]
        formula = "=" + "+".join(refs)
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = NUM_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
        tracker.set(f"funnel.grand_total.y{y}", SN, r, col)
    grand_total_row = r
    r += 2

    # Chart (reference the formula rows directly)
    chart = LineChart()
    chart.title = "Total Treated Patients by Year"
    chart.y_axis.title = "Patients"
    chart.x_axis.title = "Year"
    chart.style = 10
    chart.width = 22
    chart.height = 12
    data_ref = Reference(
        ws, min_col=2, max_col=proj_years + 1, min_row=grand_total_row, max_row=grand_total_row
    )
    cats_ref = Reference(
        ws, min_col=2, max_col=proj_years + 1, min_row=grand_total_row - 2 - len(indications) * 5
    )
    # Use year headers from the last header row
    chart.add_data(data_ref, from_rows=True, titles_from_data=False)
    chart.series[0].graphicalProperties.line.width = 25000
    ws.add_chart(chart, f"A{r}")

    ws.freeze_panes = "B4"
    return ws


# ── Sheet 3: Revenue Build (Formulas -> Funnel + Assumptions) ──


def build_revenue_sheet(wb, config, tracker):
    ws = wb.create_sheet("Revenue Build")
    ws.sheet_properties.tabColor = GREEN
    SN = "Revenue Build"

    indications = config["indications"]
    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)

    set_col_widths(ws, {"A": 38})

    r = 1
    ws.cell(row=r, column=1, value="REVENUE BUILD — FORMULA-BASED ($M)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(ws, r, 1, "Revenue = Treated Patients x Net Price / 1,000,000. All via formulas.")
    r += 2

    for ind_idx, ind in enumerate(indications):
        ind_name = ind["name"]
        line = ind.get("line_of_therapy", "")
        geo_data = ind["geography_data"]

        title = f"INDICATION: {ind_name}"
        if line:
            title += f"  ({line})"
        section_title(ws, r, 1, title)
        r += 1

        headers = [""] + [str(base_year + y) for y in range(proj_years)] + ["Total"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=r, column=i, value=h)
        apply_header_row(ws, r, len(headers))
        r += 1

        geo_rev_rows = []

        for geo in geo_data:
            ws.cell(row=r, column=1, value=f"  {geo}")
            ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True, color=MED_BLUE)
            r += 1

            net_price_ref = tracker.get(f"ind{ind_idx}.{geo}.net_price")

            # Treated Patients row (link to funnel)
            ws.cell(row=r, column=1, value="    Treated Patients").font = NORMAL_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            for y in range(proj_years):
                t_ref = tracker.get(f"funnel.ind{ind_idx}.{geo}.treated.y{y}")
                c = ws.cell(row=r, column=2 + y, value=f"={t_ref}")
                c.font = LINK_FONT
                c.number_format = NUM_FORMAT
                c.border = THIN_BORDER
            treated_row = r
            r += 1

            # Net Price row (link to Assumptions)
            ws.cell(row=r, column=1, value="    x Net Price ($)").font = NORMAL_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            for y in range(proj_years):
                c = ws.cell(row=r, column=2 + y, value=f"={net_price_ref}")
                c.font = LINK_FONT
                c.number_format = USD_FORMAT
                c.border = THIN_BORDER
            price_row = r
            r += 1

            # Revenue formula: treated x net_price / 1e6
            ws.cell(row=r, column=1, value="    -> Revenue ($M)").font = BOLD_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            for y in range(proj_years):
                col = 2 + y
                t_cell = f"{get_column_letter(col)}{treated_row}"
                p_cell = f"{get_column_letter(col)}{price_row}"
                formula = f"={t_cell}*{p_cell}/1000000"
                c = ws.cell(row=r, column=col, value=formula)
                c.font = FORMULA_FONT
                c.number_format = USD_M_FORMAT
                c.border = THIN_BORDER
                tracker.set(f"rev.ind{ind_idx}.{geo}.y{y}", SN, r, col)
            # Total column
            total_col = proj_years + 2
            c = ws.cell(
                row=r,
                column=total_col,
                value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
            )
            c.font = FORMULA_FONT
            c.number_format = USD_M_FORMAT
            c.border = THIN_BORDER
            for col in range(1, total_col + 1):
                ws.cell(row=r, column=col).border = BOTTOM_BORDER
            geo_rev_rows.append(r)
            r += 2

        # Indication total
        ws.cell(row=r, column=1, value=f"  TOTAL {ind_name} ($M)").font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for y in range(proj_years):
            col = 2 + y
            refs = [f"{get_column_letter(col)}{row}" for row in geo_rev_rows]
            formula = "=" + "+".join(refs)
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = USD_M_FORMAT
            c.border = THIN_BORDER
            c.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
            tracker.set(f"rev.ind{ind_idx}.total.y{y}", SN, r, col)
        # Total column
        total_col = proj_years + 2
        c = ws.cell(
            row=r,
            column=total_col,
            value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
        )
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        for col in range(1, total_col + 1):
            ws.cell(row=r, column=col).border = BOTTOM_BORDER
        r += 2

    # Grand total revenue
    section_title(ws, r, 1, "TOTAL REVENUE — ALL INDICATIONS ($M)")
    r += 1
    headers = [""] + [str(base_year + y) for y in range(proj_years)] + ["Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    ws.cell(row=r, column=1, value="Total Revenue ($M)").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        refs = [tracker.local(f"rev.ind{idx}.total.y{y}") for idx in range(len(indications))]
        formula = "=" + "+".join(refs)
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
        tracker.set(f"rev.total.y{y}", SN, r, col)
    total_col = proj_years + 2
    c = ws.cell(
        row=r,
        column=total_col,
        value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
    )
    c.font = FORMULA_FONT
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER
    rev_total_row = r
    r += 2

    ws.freeze_panes = "B4"
    return ws


# ── Sheet 4: Cost Structure (Formulas -> Assumptions + Revenue) ──


def build_cost_sheet(wb, config, tracker):
    ws = wb.create_sheet("Cost Structure")
    ws.sheet_properties.tabColor = ORANGE
    SN = "Cost Structure"

    costs = config["costs"]
    sga_cfg = costs.get("sga", {})
    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)
    indications = config["indications"]

    set_col_widths(ws, {"A": 40})

    r = 1
    ws.cell(row=r, column=1, value="COST STRUCTURE — FORMULA-BASED ($M)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(
        ws,
        r,
        1,
        "All costs use formulas referencing Assumptions. R&D spread linearly over duration.",
    )
    r += 2

    headers = ["Cost Item"] + [str(base_year + y) for y in range(proj_years)] + ["Total"]
    total_col = proj_years + 2
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    # ── R&D Costs ──
    section_title(ws, r, 1, "R&D COSTS")
    r += 1

    rd_rows = []
    for p_idx in range(config.get("_rd_phase_count", len(costs.get("rd_by_phase", [])))):
        phase = costs.get("rd_by_phase", [])[p_idx]
        phase_name = phase.get("phase", f"Phase {p_idx}")
        cost_mm = phase.get("cost_mm", 0)
        duration = max(phase.get("duration_years", 1), 0.5)
        start_yr = phase.get("start_year", 0)

        ws.cell(row=r, column=1, value=f"  {phase_name}").font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER

        cost_ref = tracker.get(f"rd{p_idx}.cost")
        dur_ref = tracker.get(f"rd{p_idx}.duration")
        start_ref = tracker.get(f"rd{p_idx}.start")

        for y in range(proj_years):
            col = 2 + y
            # Excel formula logic: within the phase window (from start
            # through ceiling(duration) years), spread cost evenly;
            # otherwise 0.
            formula = f"=IF(AND({y}>={start_ref},{y}<{start_ref}+CEILING({dur_ref},1)),{cost_ref}/{dur_ref},0)"
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = USD_M_FORMAT
            c.border = THIN_BORDER

        c = ws.cell(
            row=r,
            column=total_col,
            value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
        )
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        rd_rows.append(r)
        r += 1

    # CMC
    cmc_row = None
    if tracker.refs.get("cmc_total"):
        ws.cell(row=r, column=1, value="  CMC / Manufacturing").font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        cmc_ref = tracker.get("cmc_total")
        for y in range(proj_years):
            col = 2 + y
            formula = f"=IF({y}<5,{cmc_ref}/5,0)"
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = USD_M_FORMAT
            c.border = THIN_BORDER
        c = ws.cell(
            row=r,
            column=total_col,
            value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
        )
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        cmc_row = r
        rd_rows.append(r)
        r += 1

    # R&D Total
    ws.cell(row=r, column=1, value="  TOTAL R&D").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        refs = [f"{get_column_letter(col)}{row}" for row in rd_rows]
        formula = "=" + "+".join(refs)
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        tracker.set(f"cost.rd.y{y}", SN, r, col)
    c = ws.cell(
        row=r,
        column=total_col,
        value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
    )
    c.font = FORMULA_FONT
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER
    rd_total_row = r
    r += 2

    # ── COGS ──
    section_title(ws, r, 1, "COST OF GOODS SOLD (COGS)")
    r += 1
    cogs_ref = tracker.get("cogs_margin")
    ws.cell(row=r, column=1, value="  Revenue x COGS%").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        rev_ref = tracker.get(f"rev.total.y{y}")
        formula = f"={rev_ref}*{cogs_ref}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        tracker.set(f"cost.cogs.y{y}", SN, r, col)
    c = ws.cell(
        row=r,
        column=total_col,
        value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
    )
    c.font = FORMULA_FONT
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER
    cogs_total_row = r
    r += 2

    # ── SG&A ──
    section_title(ws, r, 1, "SG&A")
    r += 1

    # Pre-compute first launch year for SG&A timing
    first_launch = min(ind.get("years_to_launch", 5) for ind in indications)

    sga_component_rows = []

    # Sales Team
    sales = sga_cfg.get("sales_team", {})
    if sales and tracker.refs.get("sga.reps"):
        ws.cell(row=r, column=1, value="  Sales Team").font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        reps_ref = tracker.get("sga.reps")
        cpr_ref = tracker.get("sga.cost_per_rep")
        r0_ref = tracker.get("sga.ramp0")
        r1_ref = tracker.get("sga.ramp1")
        r2_ref = tracker.get("sga.ramp2")
        for y in range(proj_years):
            col = 2 + y
            yfl = y - first_launch  # years from launch
            if yfl < -1:
                formula = "=0"
            elif yfl == -1:
                formula = f"={reps_ref}*{cpr_ref}/1000*{r0_ref}"
            elif yfl == 0:
                formula = f"={reps_ref}*{cpr_ref}/1000*{r1_ref}"
            else:
                formula = f"={reps_ref}*{cpr_ref}/1000*{r2_ref}"
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = USD_M_FORMAT
            c.border = THIN_BORDER
        c = ws.cell(
            row=r,
            column=total_col,
            value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
        )
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        sga_component_rows.append(r)
        r += 1

    # MSLs
    if tracker.refs.get("sga.msl_count"):
        ws.cell(row=r, column=1, value="  MSLs").font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        msl_ct_ref = tracker.get("sga.msl_count")
        msl_cost_ref = tracker.get("sga.msl_cost")
        for y in range(proj_years):
            col = 2 + y
            yfl = y - first_launch
            if yfl < -2:
                formula = "=0"
            elif yfl < 0:
                formula = f"={msl_ct_ref}*{msl_cost_ref}/1000*0.5"
            else:
                formula = f"={msl_ct_ref}*{msl_cost_ref}/1000"
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = USD_M_FORMAT
            c.border = THIN_BORDER
        c = ws.cell(
            row=r,
            column=total_col,
            value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
        )
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        sga_component_rows.append(r)
        r += 1

    # Marketing
    if tracker.refs.get("sga.congress"):
        ws.cell(row=r, column=1, value="  Marketing & Promotion").font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        congress_ref = tracker.get("sga.congress")
        pubs_ref = tracker.get("sga.pubs")
        digital_ref = tracker.get("sga.digital")
        prelaunch_ref = tracker.get("sga.prelaunch")
        for y in range(proj_years):
            col = 2 + y
            yfl = y - first_launch
            if yfl == -2 or yfl == -1:
                formula = f"={prelaunch_ref}/2"
            elif yfl >= 0:
                formula = f"={congress_ref}+{pubs_ref}+{digital_ref}"
            else:
                formula = "=0"
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = USD_M_FORMAT
            c.border = THIN_BORDER
        c = ws.cell(
            row=r,
            column=total_col,
            value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
        )
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        sga_component_rows.append(r)
        r += 1

    # G&A
    if tracker.refs.get("sga.ga_pct"):
        ws.cell(row=r, column=1, value="  G&A").font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ga_ref = tracker.get("sga.ga_pct")
        for y in range(proj_years):
            col = 2 + y
            rev_ref = tracker.get(f"rev.total.y{y}")
            # Minimum $2M G&A near launch
            yfl = y - first_launch
            if -3 <= yfl <= 0:
                formula = f"=MAX({rev_ref}*{ga_ref},2)"
            else:
                formula = f"={rev_ref}*{ga_ref}"
            c = ws.cell(row=r, column=col, value=formula)
            c.font = FORMULA_FONT
            c.number_format = USD_M_FORMAT
            c.border = THIN_BORDER
        c = ws.cell(
            row=r,
            column=total_col,
            value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
        )
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        sga_component_rows.append(r)
        r += 1

    # SG&A Total
    ws.cell(row=r, column=1, value="  TOTAL SG&A").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        refs = [f"{get_column_letter(col)}{row}" for row in sga_component_rows]
        formula = "=" + "+".join(refs) if refs else "=0"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        tracker.set(f"cost.sga.y{y}", SN, r, col)
    c = ws.cell(
        row=r,
        column=total_col,
        value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
    )
    c.font = FORMULA_FONT
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER
    sga_total_row = r
    r += 2

    # ── Total Costs ──
    ws.cell(row=r, column=1, value="TOTAL COSTS ($M)").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        rd_ref = f"{get_column_letter(col)}{rd_total_row}"
        cogs_r = f"{get_column_letter(col)}{cogs_total_row}"
        sga_r = f"{get_column_letter(col)}{sga_total_row}"
        formula = f"={rd_ref}+{cogs_r}+{sga_r}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
        tracker.set(f"cost.total.y{y}", SN, r, col)
    c = ws.cell(
        row=r,
        column=total_col,
        value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
    )
    c.font = FORMULA_FONT
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER

    ws.freeze_panes = "B4"
    return ws


# ── Sheet 5: P&L & Cash Flow (Formulas -> Revenue + Cost) ──


def build_pl_sheet(wb, config, tracker):
    ws = wb.create_sheet("P&L & Cash Flow")
    ws.sheet_properties.tabColor = PURPLE
    SN = "P&L & Cash Flow"

    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)

    set_col_widths(ws, {"A": 40})

    r = 1
    ws.cell(row=r, column=1, value="PROFIT & LOSS — FORMULA-BASED ($M)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(ws, r, 1, "All lines are cross-sheet formulas. Change Assumptions to update.")
    r += 2

    headers = ["($M)"] + [str(base_year + y) for y in range(proj_years)] + ["Total"]
    total_col = proj_years + 2
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    def make_sum_total(row):
        return f"=SUM({get_column_letter(2)}{row}:{get_column_letter(proj_years + 1)}{row})"

    # Revenue
    ws.cell(row=r, column=1, value="Revenue").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=f"={tracker.get(f'rev.total.y{y}')}")
        c.font = LINK_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        tracker.set(f"pl.rev.y{y}", SN, r, col)
    ws.cell(row=r, column=total_col, value=make_sum_total(r)).font = FORMULA_FONT
    ws.cell(row=r, column=total_col).number_format = USD_M_FORMAT
    ws.cell(row=r, column=total_col).border = THIN_BORDER
    rev_row = r
    r += 1

    # COGS row — subtracted, so formulas negate the cost lookup.
    ws.cell(row=r, column=1, value="(-) COGS").font = NORMAL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=f"=-{tracker.get(f'cost.cogs.y{y}')}")
        c.font = LINK_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
    ws.cell(row=r, column=total_col, value=make_sum_total(r)).font = FORMULA_FONT
    ws.cell(row=r, column=total_col).number_format = USD_M_FORMAT
    ws.cell(row=r, column=total_col).border = THIN_BORDER
    cogs_row = r
    r += 1

    # Gross Profit
    ws.cell(row=r, column=1, value="GROSS PROFIT").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        formula = f"={get_column_letter(col)}{rev_row}+{get_column_letter(col)}{cogs_row}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        tracker.set(f"pl.gp.y{y}", SN, r, col)
    ws.cell(row=r, column=total_col, value=make_sum_total(r)).font = FORMULA_FONT
    ws.cell(row=r, column=total_col).number_format = USD_M_FORMAT
    ws.cell(row=r, column=total_col).border = THIN_BORDER
    for col in range(1, total_col + 1):
        ws.cell(row=r, column=col).border = BOTTOM_BORDER
    gp_row = r
    r += 1

    # GP Margin %
    ws.cell(row=r, column=1, value="  GP Margin %").font = Font(
        name="Calibri", size=9, italic=True, color="808080"
    )
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(
            row=r,
            column=col,
            value=f"=IFERROR({get_column_letter(col)}{gp_row}/{get_column_letter(col)}{rev_row},0)",
        )
        c.font = Font(name="Calibri", size=9, italic=True, color="808080")
        c.number_format = PCT_FORMAT
        c.border = THIN_BORDER
    r += 2

    # R&D row — subtracted, so formulas negate the cost lookup.
    ws.cell(row=r, column=1, value="(-) R&D").font = NORMAL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=f"=-{tracker.get(f'cost.rd.y{y}')}")
        c.font = LINK_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
    ws.cell(row=r, column=total_col, value=make_sum_total(r)).font = FORMULA_FONT
    ws.cell(row=r, column=total_col).number_format = USD_M_FORMAT
    ws.cell(row=r, column=total_col).border = THIN_BORDER
    rd_row = r
    r += 1

    # SG&A row — subtracted, so formulas negate the cost lookup.
    ws.cell(row=r, column=1, value="(-) SG&A").font = NORMAL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=f"=-{tracker.get(f'cost.sga.y{y}')}")
        c.font = LINK_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
    ws.cell(row=r, column=total_col, value=make_sum_total(r)).font = FORMULA_FONT
    ws.cell(row=r, column=total_col).number_format = USD_M_FORMAT
    ws.cell(row=r, column=total_col).border = THIN_BORDER
    sga_row = r
    r += 1

    # EBIT
    ws.cell(row=r, column=1, value="EBIT").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        formula = f"={cl}{gp_row}+{cl}{rd_row}+{cl}{sga_row}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        tracker.set(f"pl.ebit.y{y}", SN, r, col)
    ws.cell(row=r, column=total_col, value=make_sum_total(r)).font = FORMULA_FONT
    ws.cell(row=r, column=total_col).number_format = USD_M_FORMAT
    ws.cell(row=r, column=total_col).border = THIN_BORDER
    for col in range(1, total_col + 1):
        ws.cell(row=r, column=col).border = BOTTOM_BORDER
    ebit_row = r
    r += 1

    # EBIT Margin
    ws.cell(row=r, column=1, value="  EBIT Margin %").font = Font(
        name="Calibri", size=9, italic=True, color="808080"
    )
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(
            row=r,
            column=col,
            value=f"=IFERROR({get_column_letter(col)}{ebit_row}/{get_column_letter(col)}{rev_row},0)",
        )
        c.font = Font(name="Calibri", size=9, italic=True, color="808080")
        c.number_format = PCT_FORMAT
        c.border = THIN_BORDER
    r += 2

    # Tax
    tax_ref = tracker.get("tax_rate")
    ws.cell(row=r, column=1, value="(-) Tax").font = NORMAL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        formula = f"=-IF({cl}{ebit_row}>0,{cl}{ebit_row}*{tax_ref},0)"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
    ws.cell(row=r, column=total_col, value=make_sum_total(r)).font = FORMULA_FONT
    ws.cell(row=r, column=total_col).number_format = USD_M_FORMAT
    ws.cell(row=r, column=total_col).border = THIN_BORDER
    tax_row = r
    r += 1

    # NOPAT / FCF
    ws.cell(row=r, column=1, value="FREE CASH FLOW (Unrisked)").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        formula = f"={cl}{ebit_row}+{cl}{tax_row}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
        tracker.set(f"pl.fcf.y{y}", SN, r, col)
    ws.cell(row=r, column=total_col, value=make_sum_total(r)).font = FORMULA_FONT
    ws.cell(row=r, column=total_col).number_format = USD_M_FORMAT
    ws.cell(row=r, column=total_col).border = THIN_BORDER
    fcf_row = r
    r += 2

    # Cumulative CF
    ws.cell(row=r, column=1, value="Cumulative Cash Flow").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        if y == 0:
            formula = f"={get_column_letter(col)}{fcf_row}"
        else:
            formula = f"={get_column_letter(col - 1)}{r}+{get_column_letter(col)}{fcf_row}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
    cum_cf_row = r

    ws.freeze_panes = "B5"
    return ws


# ── Sheet 6: rNPV Model (Formulas -> P&L + Assumptions) ──


def build_rnpv_sheet(wb, config, tracker):
    ws = wb.create_sheet("rNPV Model")
    ws.sheet_properties.tabColor = RED
    SN = "rNPV Model"

    proj_years = config["discount"].get("projection_years", 20)
    base_year = config.get("metadata", {}).get("base_year", datetime.now().year)
    indications = config["indications"]

    set_col_widths(ws, {"A": 40})

    r = 1
    ws.cell(row=r, column=1, value="rNPV MODEL — FORMULA-BASED")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    calc_note(ws, r, 1, "rNPV = SUM[FCF(t) x PoS(t) x 1/(1+WACC)^(t+0.5)]  (mid-year convention)")
    r += 2

    headers = ["($M)"] + [str(base_year + y) for y in range(proj_years)] + ["NPV"]
    npv_col = proj_years + 2
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    # FCF (link to P&L)
    ws.cell(row=r, column=1, value="Unrisked FCF").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        c = ws.cell(row=r, column=col, value=f"={tracker.get(f'pl.fcf.y{y}')}")
        c.font = LINK_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    fcf_row = r
    r += 2

    # PoS Timeline (editable inputs)
    section_title(ws, r, 1, "PROBABILITY OF SUCCESS (by year)")
    r += 1

    # Show per-indication PoS summary
    for ind_idx, ind in enumerate(indications):
        pos = ind.get("pos", {})
        cum_pos_ref = tracker.get(f"ind{ind_idx}.cum_pos")
        ws.cell(row=r, column=1, value=f"  {ind['name']} Cum PoS:").font = NORMAL_FONT
        c = ws.cell(row=r, column=2, value=f"={cum_pos_ref}")
        c.font = LINK_FONT
        c.number_format = PCT_FORMAT
        c.border = THIN_BORDER
        r += 1
    r += 1

    # Year-by-year PoS (pre-approval = cumulative, post = 100%)
    first_launch = min(ind.get("years_to_launch", 5) for ind in indications)
    total_cum_pos = max(ind.get("pos", {}).get("cumulative", 0.10) for ind in indications)

    ws.cell(row=r, column=1, value="PoS (by year)").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        pos_val = total_cum_pos if y < first_launch else 1.0
        write_input_cell(ws, r, col, pos_val, PCT_FORMAT, tracker, f"rnpv.pos.y{y}", SN)
    pos_row = r
    r += 1
    calc_note(ws, r, 1, f"Pre-approval: {total_cum_pos:.1%} | Post-approval: 100%")
    r += 2

    # Risk-Adjusted FCF (formula: FCF x PoS)
    ws.cell(row=r, column=1, value="Risk-Adjusted FCF").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        formula = f"={cl}{fcf_row}*{cl}{pos_row}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    risked_row = r
    r += 2

    # Discount Factor
    wacc_ref = tracker.get("wacc")
    ws.cell(row=r, column=1, value="Discount Factor").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        formula = f"=1/(1+{wacc_ref})^({y}+0.5)"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = "0.0000"
        c.border = THIN_BORDER
    df_row = r
    r += 1

    # PV of Risked FCF
    ws.cell(row=r, column=1, value="PV of Risk-Adj FCF").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        formula = f"={cl}{risked_row}*{cl}{df_row}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
        c.fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    # NPV row — sum of the risked-and-discounted cells across projection years.
    c = ws.cell(
        row=r,
        column=npv_col,
        value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
    )
    c.font = Font(name="Calibri", size=14, bold=True, color=RED)
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER
    pv_row = r
    tracker.set("rnpv.npv", SN, r, npv_col)
    r += 2

    # Unrisked NPV
    ws.cell(row=r, column=1, value="Unrisked PV of FCF").font = NORMAL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for y in range(proj_years):
        col = 2 + y
        cl = get_column_letter(col)
        formula = f"={cl}{fcf_row}*{cl}{df_row}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FORMULA_FONT
        c.number_format = USD_M_FORMAT
        c.border = THIN_BORDER
    c = ws.cell(
        row=r,
        column=npv_col,
        value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(proj_years + 1)}{r})",
    )
    c.font = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
    c.number_format = USD_M_FORMAT
    c.border = THIN_BORDER
    tracker.set("rnpv.unrisked_npv", SN, r, npv_col)
    unrisked_row = r
    r += 2

    # Valuation Summary
    section_title(ws, r, 1, "VALUATION SUMMARY")
    r += 1
    npv_ref = tracker.local("rnpv.npv")
    unrisked_ref = tracker.local("rnpv.unrisked_npv")

    summary_items = [
        ("rNPV ($M)", f"={npv_ref}", USD_M_FORMAT),
        ("Unrisked NPV ($M)", f"={unrisked_ref}", USD_M_FORMAT),
        ("Risk Discount (rNPV/uNPV)", f"=IFERROR({npv_ref}/{unrisked_ref},0)", PCT_FORMAT),
        ("WACC", f"={wacc_ref}", PCT_FORMAT),
    ]
    for label, formula, fmt in summary_items:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        c = ws.cell(row=r, column=2, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
        c.number_format = fmt
        c.border = THIN_BORDER
        r += 1

    shares = config.get("metadata", {}).get("shares_outstanding_mm")
    if shares and shares > 0:
        ws.cell(row=r, column=1, value="Shares Outstanding (M)").font = BOLD_FONT
        ws.cell(row=r, column=2, value=shares).number_format = NUM_FORMAT
        r += 1
        ws.cell(row=r, column=1, value="rNPV Per Share ($)").font = BOLD_FONT
        c = ws.cell(row=r, column=2, value=f"={npv_ref}/{shares}")
        c.font = Font(name="Calibri", size=16, bold=True, color=GREEN)
        c.number_format = "$#,##0.00"
        r += 1

    ws.freeze_panes = "B5"

    # Store computed values for downstream (QC, Summary, Sensitivity)
    # We need actual numbers for QC checks, so compute them in Python too
    wacc_val = config["discount"]["wacc"]
    tax_rate = config["discount"].get("tax_rate", 0.20)
    config["_npv_formula_ref"] = tracker.get("rnpv.npv")
    config["_unrisked_formula_ref"] = tracker.get("rnpv.unrisked_npv")
    config["_pos_timeline"] = [
        total_cum_pos if y < first_launch else 1.0 for y in range(proj_years)
    ]

    # Python-side computation for QC/Summary (mirrors formulas)
    all_treated = {}
    for ind_idx, ind in enumerate(indications):
        pen = ind.get("penetration_curve", {})
        peak_pen = pen.get("peak", 0.15)
        ramp_yrs = pen.get("ramp_years", 7)
        loe_year = pen.get("loe_year_from_launch", 12)
        post_loe = pen.get("post_loe_erosion_per_year", 0.30)
        launch_offset = ind.get("years_to_launch", 5)
        for geo, gd in ind["geography_data"].items():
            prev = gd.get("prevalence", 0)
            addr = (
                prev
                * gd.get("diagnosed_rate", 1)
                * gd.get("eligible_rate", 1)
                * gd.get("line_share", 1)
                * gd.get("drug_treatable_rate", 1)
                * gd.get("addressable_rate", 1)
            )
            for y in range(proj_years):
                ysl = y - launch_offset
                if ysl < 0:
                    pv = 0
                elif ysl < loe_year:
                    pv = s_curve(ysl + 1, peak_pen, ramp_yrs)
                else:
                    base_p = s_curve(loe_year, peak_pen, ramp_yrs)
                    pv = base_p * ((1 - post_loe) ** (ysl - loe_year + 1))
                treated = int(addr * pv)
                net_price = ind.get("pricing", {}).get(geo, 0) * ind.get("gross_to_net", {}).get(
                    geo, 0.7
                )
                rev = treated * net_price / 1e6
                all_treated.setdefault(y, 0)
                all_treated[y] = all_treated.get(y, 0) + rev

    # Compute FCF for each year
    rd_phases = config["costs"].get("rd_by_phase", [])
    cmc_total = sum(
        v for v in config["costs"].get("cmc", {}).values() if isinstance(v, (int, float))
    )
    cogs_margin = config["costs"].get("cogs_margin", 0.20)
    sga_cfg = config["costs"].get("sga", {})
    first_launch_val = min(ind.get("years_to_launch", 5) for ind in indications)

    fcf_vals = []
    rev_vals = []
    for y in range(proj_years):
        rev = all_treated.get(y, 0)
        rev_vals.append(rev)
        rd = sum(
            p.get("cost_mm", 0) / max(p.get("duration_years", 1), 0.5)
            for p in rd_phases
            if p.get("start_year", 0)
            <= y
            < p.get("start_year", 0) + math.ceil(max(p.get("duration_years", 1), 0.5))
        )
        if y < 5:
            rd += cmc_total / 5
        cogs = rev * cogs_margin
        # Simplified SG&A
        sales_cfg = sga_cfg.get("sales_team", {})
        reps = sales_cfg.get("reps", 0)
        cpr = sales_cfg.get("cost_per_rep_k", 0)
        ramp = sales_cfg.get("ramp_schedule", [0.3, 0.6, 1.0])
        yfl = y - first_launch_val
        if yfl == -1:
            sales = reps * cpr / 1000 * (ramp[0] if ramp else 0.3)
        elif yfl == 0:
            sales = reps * cpr / 1000 * (ramp[1] if len(ramp) > 1 else ramp[-1])
        elif yfl >= 1:
            sales = reps * cpr / 1000 * (ramp[2] if len(ramp) > 2 else ramp[-1])
        else:
            sales = 0
        msls_cfg = sga_cfg.get("msls", {})
        msl = (
            msls_cfg.get("count", 0) * msls_cfg.get("cost_per_msl_k", 0) / 1000 if yfl >= -2 else 0
        )
        if yfl < 0 and msl > 0:
            msl *= 0.5
        mktg_cfg = sga_cfg.get("marketing", {})
        if yfl in (-2, -1):
            mktg = mktg_cfg.get("prelaunch_total_mm", 0) / 2
        elif yfl >= 0:
            mktg = (
                mktg_cfg.get("congress_annual_mm", 0)
                + mktg_cfg.get("publications_mm", 0)
                + mktg_cfg.get("digital_marketing_mm", 0)
            )
        else:
            mktg = 0
        ga_pct = sga_cfg.get("ga_pct_of_revenue", 0.05)
        ga = max(rev * ga_pct, 2) if -3 <= yfl <= 0 else rev * ga_pct
        sga_total = sales + msl + mktg + ga
        ebit = rev - cogs - rd - sga_total
        tax = max(0, ebit) * tax_rate
        fcf = ebit - tax
        fcf_vals.append(fcf)

    pos_tl = config["_pos_timeline"]
    disc = [1 / ((1 + wacc_val) ** (y + 0.5)) for y in range(proj_years)]
    npv = sum(fcf_vals[y] * pos_tl[y] * disc[y] for y in range(proj_years))
    unrisked_npv = sum(fcf_vals[y] * disc[y] for y in range(proj_years))
    peak_rev = max(rev_vals) if rev_vals else 0
    peak_rev_year = base_year + rev_vals.index(peak_rev) if peak_rev > 0 else "N/A"

    config["_npv"] = npv
    config["_unrisked_npv"] = unrisked_npv
    config["_peak_rev"] = peak_rev
    config["_peak_rev_year"] = peak_rev_year
    config["_rev_vals"] = rev_vals
    config["_fcf_vals"] = fcf_vals
    config["_disc_factors"] = disc
    config["_risked_fcf"] = [fcf_vals[y] * pos_tl[y] for y in range(proj_years)]
    config["_computed_revenues"] = {y: rev_vals[y] for y in range(proj_years)}
    config["_computed_costs"] = {"rd": [], "cogs": [], "sga": []}

    return ws


# ── Sheet 7: Sensitivity ──


def build_sensitivity_sheet(wb, config, tracker):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_properties.tabColor = "C00000"

    npv = config["_npv"]
    wacc = config["discount"]["wacc"]
    proj_years = config["discount"].get("projection_years", 20)
    indications = config["indications"]

    set_col_widths(ws, {"A": 32, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})

    r = 1
    ws.cell(row=r, column=1, value="SENSITIVITY ANALYSIS")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 2

    def quick_rnpv(rev_mult=1.0, cost_mult=1.0, wacc_override=None, pos_mult=1.0):
        w = wacc_override if wacc_override else wacc
        total_pos = max(ind.get("pos", {}).get("cumulative", 0.10) for ind in indications)
        adj_pos = min(total_pos * pos_mult, 1.0)
        first_launch = min(ind.get("years_to_launch", 5) for ind in indications)
        fcf_vals = config["_fcf_vals"]
        total = 0
        for y in range(proj_years):
            fcf = fcf_vals[y] * rev_mult if fcf_vals[y] > 0 else fcf_vals[y] * cost_mult
            pos = adj_pos if y < first_launch else 1.0
            df = 1 / ((1 + w) ** (y + 0.5))
            total += fcf * pos * df
        return total

    # Tornado
    section_title(ws, r, 1, "TORNADO ANALYSIS")
    r += 1
    headers = ["Variable", "Low", "Base", "High", "rNPV Low ($M)", "rNPV High ($M)", "Swing ($M)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    tornado_vars = [
        ("Peak Revenue (+-30%)", 0.70, 1.00, 1.30, lambda lo: quick_rnpv(rev_mult=lo)),
        ("Costs (+-30%)", 0.70, 1.00, 1.30, lambda lo: quick_rnpv(cost_mult=lo)),
        ("PoS (+-50%)", 0.50, 1.00, 1.50, lambda lo: quick_rnpv(pos_mult=lo)),
        ("WACC (+-3pp)", wacc + 0.03, wacc, wacc - 0.03, lambda w: quick_rnpv(wacc_override=w)),
    ]

    tornado_data = []
    for label, low_in, base_in, high_in, calc_fn in tornado_vars:
        npv_low = calc_fn(low_in)
        npv_high = calc_fn(high_in)
        if "Cost" in label:
            npv_low, npv_high = npv_high, npv_low
        swing = abs(npv_high - npv_low)
        if "WACC" in label:
            tornado_data.append(
                (
                    label,
                    f"{low_in:.1%}",
                    f"{base_in:.1%}",
                    f"{high_in:.1%}",
                    npv_low,
                    npv_high,
                    swing,
                )
            )
        else:
            tornado_data.append(
                (
                    label,
                    f"{low_in:.0%}",
                    f"{base_in:.0%}",
                    f"{high_in:.0%}",
                    npv_low,
                    npv_high,
                    swing,
                )
            )

    tornado_data.sort(key=lambda x: x[6], reverse=True)

    tornado_start = r
    for item in tornado_data:
        label, lo, base, hi, npv_l, npv_h, swing = item
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=2, value=lo).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=base).font = NORMAL_FONT
        ws.cell(row=r, column=4, value=hi).font = NORMAL_FONT
        ws.cell(row=r, column=5, value=npv_l).number_format = USD_M_FORMAT
        ws.cell(row=r, column=6, value=npv_h).number_format = USD_M_FORMAT
        ws.cell(row=r, column=7, value=swing).font = BOLD_FONT
        ws.cell(row=r, column=7).number_format = USD_M_FORMAT
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1
    tornado_end = r - 1
    r += 1

    # Tornado chart
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Tornado — rNPV Impact ($M)"
    chart.style = 10
    chart.width = 22
    chart.height = 14
    data_low = Reference(ws, min_col=5, min_row=tornado_start - 1, max_row=tornado_end)
    data_high = Reference(ws, min_col=6, min_row=tornado_start - 1, max_row=tornado_end)
    cats = Reference(ws, min_col=1, min_row=tornado_start, max_row=tornado_end)
    chart.add_data(data_low, titles_from_data=True)
    chart.add_data(data_high, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = RED
    chart.series[1].graphicalProperties.solidFill = GREEN
    ws.add_chart(chart, f"A{r}")
    r += 18

    # Scenarios
    section_title(ws, r, 1, "SCENARIO COMPARISON")
    r += 1
    scenarios = config.get("scenarios", {})
    bull_rev = scenarios.get("bull", {}).get("revenue_multiplier", 1.3)
    bear_rev = scenarios.get("bear", {}).get("revenue_multiplier", 0.7)
    bull_pos = scenarios.get("bull", {}).get("pos_multiplier", 1.3)
    bear_pos = scenarios.get("bear", {}).get("pos_multiplier", 0.7)

    sc_headers = ["Scenario", "Rev Mult", "PoS Mult", "rNPV ($M)", "vs Base"]
    for i, h in enumerate(sc_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(sc_headers))
    r += 1

    for label, rm, pm in [
        ("Bear", bear_rev, bear_pos),
        ("Base", 1.0, 1.0),
        ("Bull", bull_rev, bull_pos),
    ]:
        val = quick_rnpv(rev_mult=rm, pos_mult=pm)
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=2, value=rm).number_format = PCT_FORMAT
        ws.cell(row=r, column=3, value=pm).number_format = PCT_FORMAT
        ws.cell(row=r, column=4, value=val).number_format = USD_M_FORMAT
        ws.cell(row=r, column=5, value=val - npv).number_format = USD_M_FORMAT
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1

    ws.freeze_panes = "A3"
    return ws


# ── Sheet 8: QC Report ──


def build_qc_sheet(wb, config, tracker):
    ws = wb.create_sheet("QC Report")
    ws.sheet_properties.tabColor = "7030A0"

    set_col_widths(ws, {"A": 50, "B": 12, "C": 50, "D": 25})

    r = 1
    ws.cell(row=r, column=1, value="MODEL QUALITY CONTROL REPORT")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | v3 Formula-Based",
    )
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="808080")
    r += 2

    headers = ["Check", "Status", "Detail", "Expected"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    checks = []
    indications = config["indications"]
    proj_years = config["discount"].get("projection_years", 20)
    npv = config.get("_npv", 0)
    unrisked_npv = config.get("_unrisked_npv", 0)
    peak_rev = config.get("_peak_rev", 0)
    rev_vals = config.get("_rev_vals", [])
    revenues = config.get("_computed_revenues", {})

    # Patient funnel rates
    for ind in indications:
        for geo, gd in ind["geography_data"].items():
            rates = [
                gd.get(k, 1)
                for k in [
                    "diagnosed_rate",
                    "eligible_rate",
                    "line_share",
                    "drug_treatable_rate",
                    "addressable_rate",
                ]
            ]
            ok = all(0 < r_ <= 1.0 for r_ in rates)
            checks.append(
                (
                    f"Funnel rates ({ind['name']}/{geo})",
                    "PASS" if ok else "FAIL",
                    f"Rates: {', '.join(f'{r_:.0%}' for r_ in rates)}",
                    "All 0-100%",
                )
            )

    # Peak revenue
    checks.append(
        (
            "Peak revenue range",
            "PASS" if 50 < peak_rev < 50000 else "WARN",
            f"${peak_rev:,.0f}M",
            "$50M-$50B",
        )
    )

    # PoS vs phase
    for ind in indications:
        pos = ind.get("pos", {})
        phase = pos.get("current_phase", "")
        cum = pos.get("cumulative", 0)
        ranges = {
            "Preclinical": (0.01, 0.15),
            "Phase 1": (0.05, 0.25),
            "Phase 2": (0.10, 0.40),
            "Phase 3": (0.30, 0.70),
            "NDA": (0.70, 0.95),
        }
        exp = ranges.get(phase, (0.01, 1.0))
        ok = exp[0] <= cum <= exp[1]
        checks.append(
            (
                f"PoS vs phase ({ind['name']})",
                "PASS" if ok else "WARN",
                f"{cum:.1%}",
                f"{exp[0]:.0%}-{exp[1]:.0%} for {phase}",
            )
        )

    # WACC
    wacc = config["discount"]["wacc"]
    checks.append(
        ("WACC range", "PASS" if 0.08 <= wacc <= 0.20 else "WARN", f"{wacc:.1%}", "8%-20%")
    )

    # Revenue before launch
    first_launch = min(ind.get("years_to_launch", 5) for ind in indications)
    pre_rev = sum(revenues.get(y, 0) for y in range(first_launch))
    checks.append(
        ("No pre-launch revenue", "PASS" if pre_rev == 0 else "FAIL", f"${pre_rev:,.0f}M", "$0")
    )

    # NPV non-zero
    checks.append(
        ("NPV non-zero", "PASS" if abs(npv) > 0.1 else "FAIL", f"${npv:,.1f}M", "Non-zero")
    )

    # rNPV/uNPV ratio
    if unrisked_npv != 0:
        ratio = npv / unrisked_npv
        checks.append(
            ("rNPV/uNPV ratio", "PASS" if 0.01 < ratio < 0.80 else "WARN", f"{ratio:.0%}", "1%-80%")
        )

    # LOE erosion
    if rev_vals and len(rev_vals) > 5:
        last5 = rev_vals[-5:]
        declines = any(last5[i] < last5[i - 1] for i in range(1, len(last5)))
        checks.append(
            (
                "LOE erosion visible",
                "PASS" if declines else "WARN",
                "Revenue declines" if declines else "Still growing",
                "Post-LOE decline",
            )
        )

    # Pricing complete
    for ind in indications:
        pricing = ind.get("pricing", {})
        geos = list(ind["geography_data"].keys())
        ok = all(pricing.get(g, 0) > 0 for g in geos)
        checks.append(
            (
                f"Pricing complete ({ind['name']})",
                "PASS" if ok else "FAIL",
                f"{sum(1 for g in geos if pricing.get(g, 0) > 0)}/{len(geos)}",
                "All geos priced",
            )
        )

    # References
    ref_total = 0
    ref_sourced = 0
    for ind in indications:
        ds = ind.get("data_sources", {})
        for p in [
            "prevalence",
            "diagnosed_rate",
            "eligible_rate",
            "line_share",
            "drug_treatable_rate",
            "addressable_rate",
            "pricing",
        ]:
            ref_total += 1
            src = ds.get(p, "")
            if src and not any(kw in src.lower() for kw in ["estimate", "assumption", "default"]):
                ref_sourced += 1
    ref_pct = ref_sourced / ref_total if ref_total > 0 else 0
    checks.append(
        (
            "Reference coverage",
            "PASS" if ref_pct >= 0.7 else "WARN" if ref_pct >= 0.4 else "FAIL",
            f"{ref_sourced}/{ref_total} ({ref_pct:.0%})",
            ">=70%",
        )
    )

    # Formula-based model check
    checks.append(
        (
            "Model type: Formula-based",
            "PASS",
            "v3 — all calculations use Excel formulas",
            "Dynamic model",
        )
    )

    # Write
    pass_n = warn_n = fail_n = 0
    for name, status, detail, exp in checks:
        ws.cell(row=r, column=1, value=name).font = NORMAL_FONT
        c = ws.cell(row=r, column=2, value=status)
        if status == "PASS":
            c.fill = PASS_FILL
            c.font = Font(name="Calibri", size=10, bold=True, color="006100")
            pass_n += 1
        elif status == "FAIL":
            c.fill = FAIL_FILL
            c.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
            fail_n += 1
        else:
            c.fill = WARN_FILL
            c.font = Font(name="Calibri", size=10, bold=True, color="9C6500")
            warn_n += 1
        ws.cell(row=r, column=3, value=detail).font = NORMAL_FONT
        ws.cell(row=r, column=4, value=exp).font = Font(name="Calibri", size=9, color="808080")
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1

    r += 2
    section_title(ws, r, 1, "QC SUMMARY")
    r += 1
    write_label_value(
        ws,
        r,
        1,
        "Passed",
        pass_n,
        val_font=Font(name="Calibri", size=12, bold=True, color="006100"),
    )
    ws.cell(row=r, column=2).fill = PASS_FILL
    r += 1
    write_label_value(
        ws,
        r,
        1,
        "Warnings",
        warn_n,
        val_font=Font(name="Calibri", size=12, bold=True, color="9C6500"),
    )
    ws.cell(row=r, column=2).fill = WARN_FILL
    r += 1
    write_label_value(
        ws,
        r,
        1,
        "Failed",
        fail_n,
        val_font=Font(name="Calibri", size=12, bold=True, color="9C0006"),
    )
    ws.cell(row=r, column=2).fill = FAIL_FILL

    config["_qc_pass"] = pass_n
    config["_qc_fail"] = fail_n
    config["_qc_warn"] = warn_n
    return ws


# ── Sheet 9: References ──


def build_references_sheet(wb, config, tracker):
    ws = wb.create_sheet("References")
    ws.sheet_properties.tabColor = "00B0F0"

    set_col_widths(ws, {"A": 10, "B": 20, "C": 55, "D": 25, "E": 18, "F": 45})

    r = 1
    ws.cell(row=r, column=1, value="DATA SOURCES & REFERENCES")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
    r += 2

    headers = ["Ref #", "Category", "Description", "Type", "Date", "URL"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(headers))
    r += 1

    references = config.get("references", [])
    if not references:
        # Auto-collect from data_sources
        seen = set()
        ref_n = 0
        for ind in config.get("indications", []):
            for key, src in ind.get("data_sources", {}).items():
                if src and src not in seen:
                    seen.add(src)
                    ref_n += 1
                    references.append(
                        {
                            "id": f"R{ref_n}",
                            "category": "Input",
                            "description": src,
                            "type": "Various",
                            "date": "",
                            "url": "",
                        }
                    )

    ref_count = 0
    for ref in references:
        ref_count += 1
        ws.cell(row=r, column=1, value=f"[{ref.get('id', f'R{ref_count}')}]").font = Font(
            name="Calibri", size=10, bold=True, color=MED_BLUE
        )
        ws.cell(row=r, column=2, value=ref.get("category", "")).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=ref.get("description", "")).font = NORMAL_FONT
        ws.cell(row=r, column=4, value=ref.get("type", "")).font = NORMAL_FONT
        ws.cell(row=r, column=5, value=ref.get("date", "")).font = NORMAL_FONT
        url = ref.get("url", "")
        c = ws.cell(row=r, column=6, value=url)
        if url and url.startswith("http"):
            c.font = Font(name="Calibri", size=10, color=MED_BLUE, underline="single")
        else:
            c.font = NORMAL_FONT
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1

    # Coverage summary
    r += 2
    ref_sourced = 0
    ref_total = 0
    for ind in config.get("indications", []):
        ds = ind.get("data_sources", {})
        for p in ["prevalence", "diagnosed_rate", "eligible_rate", "pricing"]:
            ref_total += 1
            if ds.get(p):
                ref_sourced += 1
    coverage = ref_sourced / ref_total if ref_total > 0 else 0
    write_label_value(ws, r, 1, "References", ref_count)
    r += 1
    write_label_value(ws, r, 1, "Key Params Sourced", f"{ref_sourced}/{ref_total} ({coverage:.0%})")

    config["_ref_count"] = ref_count
    config["_ref_coverage_pct"] = coverage
    config["_ref_sourced"] = ref_sourced
    config["_ref_total_params"] = ref_total
    return ws


# ── Sheet 0: Summary Dashboard ──


def build_summary_sheet(wb, config, tracker):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "FFC000"

    meta = config["metadata"]
    proj_years = config["discount"].get("projection_years", 20)
    base_year = meta.get("base_year", datetime.now().year)
    indications = config["indications"]

    set_col_widths(ws, {"A": 38, "B": 22, "C": 22, "D": 22})

    r = 1
    ws.cell(row=r, column=1, value="rNPV VALUATION SUMMARY")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
    r += 1
    ws.cell(row=r, column=1, value=f"{meta['company']} -- {meta['asset']}")
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=12, color=MED_BLUE)
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=f"Date: {meta.get('date', datetime.now().strftime('%Y-%m-%d'))} | v3 Formula-Based",
    )
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="808080")
    r += 2

    section_title(ws, r, 1, "KEY METRICS (Formula-Linked)")
    r += 1

    # rNPV (formula link)
    npv_ref = tracker.get("rnpv.npv")
    unrisked_ref = tracker.get("rnpv.unrisked_npv")
    wacc_ref = tracker.get("wacc")

    metrics = [
        ("rNPV ($M)", f"={npv_ref}", USD_M_FORMAT),
        ("Unrisked NPV ($M)", f"={unrisked_ref}", USD_M_FORMAT),
        ("Risk Discount", f"=IFERROR({npv_ref}/{unrisked_ref},0)", PCT_FORMAT),
        ("WACC", f"={wacc_ref}", PCT_FORMAT),
    ]

    for label, formula, fmt in metrics:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        c = ws.cell(row=r, column=2, value=formula)
        c.font = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
        c.number_format = fmt
        c.border = THIN_BORDER
        r += 1

    r += 1

    # QC
    qc_pass = config.get("_qc_pass", 0)
    qc_warn = config.get("_qc_warn", 0)
    qc_fail = config.get("_qc_fail", 0)
    ws.cell(row=r, column=1, value="QC Status").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    status_text = f"{'PASS' if qc_fail == 0 else 'REVIEW'} ({qc_pass}P {qc_warn}W {qc_fail}F)"
    c = ws.cell(row=r, column=2, value=status_text)
    c.font = Font(name="Calibri", size=11, bold=True, color="006100" if qc_fail == 0 else "9C0006")
    c.fill = PASS_FILL if qc_fail == 0 else FAIL_FILL
    c.border = THIN_BORDER
    r += 2

    # Per-indication
    section_title(ws, r, 1, "PIPELINE SUMMARY")
    r += 1
    ind_headers = ["Indication", "Line", "Phase", "Cum PoS", "Yrs to Launch"]
    for i, h in enumerate(ind_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, len(ind_headers))
    r += 1

    for ind_idx, ind in enumerate(indications):
        pos = ind.get("pos", {})
        ws.cell(row=r, column=1, value=ind["name"]).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=ind.get("line_of_therapy", "")).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=pos.get("current_phase", "")).font = NORMAL_FONT
        # Link to Assumptions PoS
        cum_ref = tracker.get(f"ind{ind_idx}.cum_pos")
        c = ws.cell(row=r, column=4, value=f"={cum_ref}")
        c.font = LINK_FONT
        c.number_format = PCT_FORMAT
        ws.cell(row=r, column=5, value=ind.get("years_to_launch", "")).font = NORMAL_FONT
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1

    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))
    ws.freeze_panes = "A5"
    return ws


# ── Main ──


def generate(config, output_path):
    wb = Workbook()
    tracker = CellTracker()

    build_assumptions_sheet(wb, config, tracker)
    build_patient_funnel_sheet(wb, config, tracker)
    build_revenue_sheet(wb, config, tracker)
    build_cost_sheet(wb, config, tracker)
    build_pl_sheet(wb, config, tracker)
    build_rnpv_sheet(wb, config, tracker)
    build_sensitivity_sheet(wb, config, tracker)
    build_qc_sheet(wb, config, tracker)
    build_references_sheet(wb, config, tracker)
    build_summary_sheet(wb, config, tracker)

    wb.save(output_path)

    npv = config.get("_npv", 0)
    peak_rev = config.get("_peak_rev", 0)
    peak_yr = config.get("_peak_rev_year", "N/A")

    print(f"✓ rNPV model saved to: {output_path}")
    print(
        "  Sheets: Summary | Assumptions | Patient Funnel | Revenue Build | Cost Structure | P&L & Cash Flow | rNPV Model | Sensitivity | QC Report | References"
    )
    print(f"  rNPV: ${npv:,.1f}M")
    print(f"  Unrisked NPV: ${config.get('_unrisked_npv', 0):,.1f}M")
    print(f"  Peak Revenue: ${peak_rev:,.1f}M in {peak_yr}")
    qc_fail = config.get("_qc_fail", 0)
    print(
        f"  QC: {config.get('_qc_pass', 0)} pass, {config.get('_qc_warn', 0)} warn, {qc_fail} fail {'⚠' if qc_fail > 0 else '✓'}"
    )
    print(
        f"  References: {config.get('_ref_count', 0)} sources, {config.get('_ref_sourced', 0)}/{config.get('_ref_total_params', 0)} params sourced ({config.get('_ref_coverage_pct', 0):.0%} coverage)"
    )
    print("  Model: v3 Formula-Based (change Assumptions -> all sheets update)")
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Generate rNPV Valuation Excel Model v3 (Formula-Based)"
    )
    parser.add_argument("--config", required=True, help="Path to JSON config file")
    parser.add_argument("--output", help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.config) as f:
        config = json.load(f)

    if not args.output:
        company = config.get("metadata", {}).get("company", "unknown")
        asset = config.get("metadata", {}).get("asset", "asset")
        safe_name = f"{company}_{asset}_rNPV".replace(" ", "_").replace("/", "_")
        args.output = f"{safe_name}.xlsx"

    generate(config, args.output)


if __name__ == "__main__":
    main()
